import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def create_output_folder(folder_path):
    """Create output folder if it doesn't exist"""
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
        print(f"Created folder: {folder_path}")

def read_json_file(file_path):
    """Read and parse JSON file"""
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            data = json.load(file)
        print(f"Successfully read JSON file: {file_path}")
        return data
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
        return None
    except json.JSONDecodeError as e:
        print(f"Error: Invalid JSON format in '{file_path}': {e}")
        return None
    except Exception as e:
        print(f"Error reading file '{file_path}': {e}")
        return None

def create_financial_table_sheet(workbook, sheet_name, note_data):
    """Create a properly formatted financial table sheet"""
    ws = workbook.create_sheet(title=sheet_name)
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    right_alignment = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    current_row = 1
    
    # Add Note Title
    note_title = note_data.get('full_title', note_data.get('note_title', 'Note'))
    ws.cell(row=current_row, column=1, value=note_title)
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
    current_row += 2
    
    # Process table_data if available
    if 'table_data' in note_data and note_data['table_data']:
        table_data = note_data['table_data']
        
        # Create DataFrame from table_data
        df = pd.DataFrame(table_data)
        
        # Add table headers
        for col_num, column_name in enumerate(df.columns, 1):
            cell = ws.cell(row=current_row, column=col_num, value=column_name.replace('_', ' ').title())
            cell.font = header_font
            cell.border = border
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        
        # Process structure
        for section in json_data.get('structure', []):
            category = section.get('category', '')
            subcategories = section.get('subcategories', [])
            
            # Add category header if exists
            if category:
                ws.merge_cells(f'A{row}:C{row}')
                cell = ws[f'A{row}']
                cell.value = category
                cell.font = Font(bold=True, size=10)
                cell.border = border
                cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                row += 1
            
            # Add subcategories
            for subcat in subcategories:
                label = subcat.get('label', '')
                current_value = subcat.get('value', '0.00')
                previous_value = subcat.get('previous_value', '0.00')
                
                # Handle special cases for totals
                if label.startswith('March 31'):
                    continue  # Skip these as they're handled in headers
                
                # Format values
                if isinstance(current_value, str) and current_value.startswith('{'):
                    current_formatted = "-"
                else:
                    current_formatted = self.format_currency(self.parse_value(current_value))
                
                if isinstance(previous_value, str) and previous_value.startswith('{'):
                    previous_formatted = "-"
                else:
                    previous_formatted = self.format_currency(self.parse_value(previous_value))
                
                # Add row
                ws.cell(row=row, column=1, value=label).font = normal_font
                ws.cell(row=row, column=2, value=current_formatted).font = normal_font
                ws.cell(row=row, column=3, value=previous_formatted).font = normal_font
                
                # Apply borders
                for col in range(1, 4):
                    ws.cell(row=row, column=col).border = border
                
                # Right align numbers
                ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
                ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
                
                row += 1
            
            # Add total if exists
            if 'total' in section:
                total_current = section.get('total', '0.00')
                total_previous = section.get('previous_total', '0.00')
                
                # Format totals
                if isinstance(total_current, str) and total_current.startswith('{'):
                    total_current_formatted = "-"
                else:
                    total_current_formatted = self.format_currency(self.parse_value(total_current))
                
                if isinstance(total_previous, str) and total_previous.startswith('{'):
                    total_previous_formatted = "-"
                else:
                    total_previous_formatted = self.format_currency(self.parse_value(total_previous))
                
                # Add total row
                ws.cell(row=row, column=1, value="Total").font = Font(bold=True, size=10)
                ws.cell(row=row, column=2, value=total_current_formatted).font = Font(bold=True, size=10)
                ws.cell(row=row, column=3, value=total_previous_formatted).font = Font(bold=True, size=10)
                
                # Apply borders and alignment
                for col in range(1, 4):
                    ws.cell(row=row, column=col).border = border
                
                ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
                ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
                
                row += 1
        
        # Add notes and disclosures
        if 'notes_and_disclosures' in json_data:
            row += 1
            for note in json_data['notes_and_disclosures']:
                ws.merge_cells(f'A{row}:C{row}')
                ws[f'A{row}'] = note
                ws[f'A{row}'].font = Font(size=9, italic=True)
                row += 1
        
        # Auto-adjust column widths
        for col in range(1, 4):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save the file
        wb.save(output_file)
        print(f"Excel file created: {output_file}")
    
    def process_all_files(self):
        """Process all JSON files in the input folder"""
        if not os.path.exists(self.input_folder):
            print(f"Input folder '{self.input_folder}' does not exist!")
            return
        
        json_files = [f for f in os.listdir(self.input_folder) if f.endswith('.json')]
        
        if not json_files:
            print(f"No JSON files found in '{self.input_folder}'")
            return
        
        for json_file in json_files:
            input_path = os.path.join(self.input_folder, json_file)
            output_filename = json_file.replace('.json', '.xlsx')
            output_path = os.path.join(self.output_folder, output_filename)
            
            print(f"Processing {json_file}...")
            
            json_data = self.load_json_data(input_path)
            if json_data:
                self.create_excel_from_json(json_data, output_path)
            else:
                print(f"Failed to process {json_file}")
    
    def process_single_file(self, input_file, output_file=None):
        """Process a single JSON file"""
        if not output_file:
            output_file = input_file.replace('.json', '.xlsx')
            output_file = os.path.join(self.output_folder, os.path.basename(output_file))
        
        json_data = self.load_json_data(input_file)
        if json_data:
            self.create_excel_from_json(json_data, output_file)
        else:
            print(f"Failed to process {input_file}")

def main():
    # Example usage
    converter = FinancialNotesConverter()
    
    # Process all files in the generated_notes folder
    converter.process_all_files()
    
    # Or process a single file
    # converter.process_single_file('generated_notes/paste.txt', 'output.xlsx')

if __name__ == "__main__":
    main()
