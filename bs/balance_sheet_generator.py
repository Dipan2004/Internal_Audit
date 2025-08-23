"""
Enhanced Balance Sheet Generator
Production-ready script for extracting, analyzing, and generating Excel balance sheets from financial JSON data.
"""
import os
import json
import re
import logging
from datetime import datetime
from typing import Any, Dict, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
import requests
from dotenv import load_dotenv
from pydantic import BaseModel, Field, ValidationError
from pydantic_settings import BaseSettings

# Import the template handler
from balance_sheet_template_handler import BalanceSheetTemplate, STANDARD_NOTES_MAPPING

load_dotenv()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s %(levelname)s %(name)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

class Settings(BaseSettings):
    """Application settings loaded from environment variables or .env file."""
    api_key: str = Field(default_factory=lambda: os.getenv("OPENROUTER_API_KEY", ""), env="OPENROUTER_API_KEY")
    input_file: str = Field(default="data/clean_financial_data_bs.json", env="INPUT_FILE")
    output_dir: str = Field(default="data/output", env="BL_OUTPUT_DIR")

settings = Settings()

class BalanceSheetItem(BaseModel):
    category: str
    subcategory: Optional[str] = ""
    name: str
    note: Optional[str] = ""
    value_2024: float
    value_2023: float

class BalanceSheetTotals(BaseModel):
    shareholders_funds_2024: float = 0.0
    shareholders_funds_2023: float = 0.0
    share_application_money_2024: float = 0.0
    share_application_money_2023: float = 0.0
    non_current_liabilities_2024: float = 0.0
    non_current_liabilities_2023: float = 0.0
    current_liabilities_2024: float = 0.0
    current_liabilities_2023: float = 0.0
    non_current_assets_2024: float = 0.0
    non_current_assets_2023: float = 0.0
    current_assets_2024: float = 0.0
    current_assets_2023: float = 0.0
    total_equity_liabilities_2024: float = 0.0
    total_equity_liabilities_2023: float = 0.0
    total_assets_2024: float = 0.0
    total_assets_2023: float = 0.0
    balance_difference_2024: float = 0.0
    balance_difference_2023: float = 0.0

class EnhancedBalanceSheetGenerator:
    def __init__(self, api_key: str):
        self.api_key = api_key
        self.base_url = "https://openrouter.ai/api/v1/chat/completions"
        
        # Initialize template
        self.template = BalanceSheetTemplate()
        self.field_mappings = self.template.get_field_mappings()
        self.formatting_rules = self.template.get_formatting_rules()
        
        logger.info(f"Loaded template with {len(self.template.get_template_structure())} items")

    def safe_float(self, value: Any) -> float:
        """Convert various value formats to float."""
        if not value or str(value).strip() in ['-', '--', 'None', '', 'null']:
            return 0.0
        
        # Handle strings
        if isinstance(value, str):
            # Remove currency symbols and brackets
            cleaned = re.sub(r'[â‚¹,Rs\.\s\(\)]', '', value)
            # Handle negative values in brackets
            if '(' in str(value) and ')' in str(value):
                cleaned = '-' + cleaned.replace('(', '').replace(')', '')
            try:
                return float(cleaned)
            except Exception:
                return 0.0
        
        # Handle numeric values
        try:
            return float(value)
        except Exception:
            return 0.0

    def get_value_flexible(self, data: Any, date_key_2024: str = "2024-03-31 00:00:00", date_key_2023: str = "2023-03-31 00:00:00") -> tuple[float, float]:
        """Flexibly extract values from either list or dictionary format."""
        if isinstance(data, dict):
            # Dictionary format - extract by date keys
            val_2024 = self.safe_float(data.get(date_key_2024, 0))
            val_2023 = self.safe_float(data.get(date_key_2023, 0))
            return val_2024, val_2023
        
        elif isinstance(data, list):
            # List format - assume first element is 2024, second is 2023
            val_2024 = self.safe_float(data[0]) if len(data) > 0 else 0.0
            val_2023 = self.safe_float(data[1]) if len(data) > 1 else 0.0
            return val_2024, val_2023
        
        else:
            # Single value or other format
            val = self.safe_float(data)
            return val, 0.0  # Assume it's 2024 value, 2023 is 0

    def call_ai_for_analysis(self, data_summary: str) -> Dict[str, Any]:
        """Use AI to analyze and extract balance sheet data using the template structure"""
        
        template_items = self.template.get_template_structure()
        
        prompt = f"""
You are a financial analyst. Extract balance sheet data from the following JSON data and create a properly structured balance sheet.

CRITICAL REQUIREMENTS:
1. Extract ALL line items with their 2024 and 2023 values
2. Use the EXACT template structure provided below
3. Calculate missing totals where needed
4. Ensure the balance sheet balances (Assets = Equity + Liabilities)
5. Return ONLY valid JSON in the exact format specified below

Expected Balance Sheet Structure (use this EXACT structure):
{json.dumps(template_items, indent=2)}

Data to analyze:
{data_summary}

Return ONLY this JSON format:
{{
  "balance_sheet_items": [
    {{
      "category": "Shareholders' funds",
      "subcategory": "",
      "name": "Share capital",
      "note": "2",
      "value_2024": 542.52,
      "value_2023": 542.52
    }},
    {{
      "category": "Shareholders' funds", 
      "subcategory": "",
      "name": "Reserves and surplus",
      "note": "3",
      "value_2024": 3152.39,
      "value_2023": 2642.87
    }}
  ],
  "totals": {{
    "shareholders_funds_2024": 3694.91,
    "shareholders_funds_2023": 3185.39,
    "total_equity_liabilities_2024": 5246.10,
    "total_equity_liabilities_2023": 4725.23,
    "total_assets_2024": 5246.10,
    "total_assets_2023": 4725.23
  }}
}}
"""

        headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json"
        }
        
        payload = {
            "model": "anthropic/claude-3.5-sonnet",
            "messages": [{"role": "user", "content": prompt}],
            "temperature": 0.1,
            "max_tokens": 4000
        }
        
        try:
            response = requests.post(self.base_url, headers=headers, json=payload, timeout=60)
            content = response.json()['choices'][0]['message']['content']
            
            # Clean the response
            content = re.sub(r'```(?:json)?\s*', '', content).strip('`').strip()
            
            return json.loads(content)
        except Exception as e:
            logger.error(f"AI analysis failed: {e}")
            return {"balance_sheet_items": [], "totals": {}}

    def extract_from_json_structure(self, json_data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Extract data using template structure with flexible JSON mapping"""
        items = []
        
        company_data = json_data.get("company_financial_data", {})
        
        # Get template structure to guide extraction
        template_items = self.template.get_template_structure()
        
        # Extract based on template structure
        for template_item in template_items:
            item_name = template_item["name"]
            category = template_item["category"]
            subcategory = template_item.get("subcategory", "")
            note = template_item["note"]
            
            val_2024 = val_2023 = 0.0
            
            # Map template items to JSON data extraction logic
            if item_name == "Share capital":
                share_capital = company_data.get("share_capital", {})
                total_share_capital = share_capital.get("Total issued, subscribed and fully paid-up share capital", {})
                if total_share_capital:
                    val_2024, val_2023 = self.get_value_flexible(total_share_capital)
            
            elif item_name == "Reserves and surplus":
                reserves = company_data.get("reserves_and_surplus", {})
                closing_balance = reserves.get("Balance, at the end of the year", {})
                if closing_balance:
                    val_2024, val_2023 = self.get_value_flexible(closing_balance)
            
            elif item_name == "Money received against share warrants":
                # Try to find this in the data - might be in share capital or other sections
                pass  # Will be 0 if not found
            
            elif item_name == "Share application money pending allotment":
                # Try to find this in the data
                pass  # Will be 0 if not found
            
            elif item_name == "Long-term borrowings":
                borrowings = company_data.get("borrowings", {}).get("4. Long-Term Borrowings", {})
                for key, value in borrowings.items():
                    if key != "_metadata" and value is not None:
                        v24, v23 = self.get_value_flexible(value)
                        val_2024 += v24
                        val_2023 += v23
            
            elif item_name == "Deferred tax liabilities (Net)":
                deferred_tax = company_data.get("other_data", {}).get("5. Deferred Tax Liability / (Asset)", {})
                if deferred_tax:
                    dtl = deferred_tax.get("Deferred tax liability", {})
                    val_2024, val_2023 = self.get_value_flexible(dtl)
            
            elif item_name == "Short-term borrowings":
                # Look for short term borrowings in current liabilities
                current_liabilities = company_data.get("current_liabilities", {})
                # This might be in various sections, keep as 0 for now
                pass
            
            elif "total outstanding dues of micro enterprises" in item_name:
                # Extract from trade payables - micro enterprises
                current_liabilities = company_data.get("current_liabilities", {})
                trade_payables = current_liabilities.get("6. Trade Payables", {})
                # Look for micro enterprises specific data
                pass
            
            elif "total outstanding dues of creditors other than micro" in item_name:
                # Extract from trade payables - others
                current_liabilities = company_data.get("current_liabilities", {})
                trade_payables = current_liabilities.get("6. Trade Payables", {})
                for key, value in trade_payables.items():
                    if key not in ["_metadata", "Particulars", "Disputed dues"] and value is not None:
                        v24, v23 = self.get_value_flexible(value)
                        val_2024 += v24
                        val_2023 += v23
            
            elif item_name == "Other current liabilities":
                current_liabilities = company_data.get("current_liabilities", {})
                other_cl = current_liabilities.get("7. Other Current Liabilities", {})
                for key, value in other_cl.items():
                    if key != "_metadata" and value is not None:
                        v24, v23 = self.get_value_flexible(value)
                        val_2024 += v24
                        val_2023 += v23
            
            elif item_name == "Short-term provisions":
                current_liabilities = company_data.get("current_liabilities", {})
                provisions = current_liabilities.get("8. Short Term Provisions", {})
                for key, value in provisions.items():
                    if key != "_metadata" and value is not None:
                        v24, v23 = self.get_value_flexible(value)
                        val_2024 += v24
                        val_2023 += v23
            
            elif item_name == "Tangible assets":
                fixed_assets = company_data.get("fixed_assets", {})
                tangible = fixed_assets.get("tangible_assets", {}).get("", {})
                if tangible:
                    net_carrying = tangible.get("net_carrying_value", {})
                    if net_carrying:
                        if isinstance(net_carrying, dict):
                            val_2024 = self.safe_float(net_carrying.get("closing", 0))
                            val_2023 = self.safe_float(net_carrying.get("opening", 0))
                        else:
                            val_2024, val_2023 = self.get_value_flexible(net_carrying)
            
            elif item_name == "Intangible assets":
                fixed_assets = company_data.get("fixed_assets", {})
                intangible = fixed_assets.get("intangible_assets", {}).get("", {})
                if intangible:
                    net_carrying = intangible.get("net_carrying_value", {})
                    if net_carrying:
                        if isinstance(net_carrying, dict):
                            val_2024 = self.safe_float(net_carrying.get("closing", 0))
                            val_2023 = self.safe_float(net_carrying.get("opening", 0))
                        else:
                            val_2024, val_2023 = self.get_value_flexible(net_carrying)
            
            elif item_name == "Long-term loans and advances":
                lt_loans = company_data.get("loans_and_advances", {}).get("10. Long Term Loans and advances", {})
                for key, value in lt_loans.items():
                    if key != "_metadata" and value is not None:
                        v24, v23 = self.get_value_flexible(value)
                        val_2024 += v24
                        val_2023 += v23
            
            elif item_name == "Inventories":
                current_assets = company_data.get("current_assets", {})
                inventories = current_assets.get("11. Inventories", {})
                for key, value in inventories.items():
                    if key != "_metadata" and value is not None:
                        v24, v23 = self.get_value_flexible(value)
                        val_2024 += v24
                        val_2023 += v23
            
            elif item_name == "Trade receivables":
                current_assets = company_data.get("current_assets", {})
                trade_recv = current_assets.get("12. Trade receivables", {})
                for key, value in trade_recv.items():
                    if key not in ["_metadata", "Particulars", "trade_receivables_aging"] and value is not None:
                        v24, v23 = self.get_value_flexible(value)
                        val_2024 += v24
                        val_2023 += v23
            
            elif item_name == "Cash and cash equivalents":
                current_assets = company_data.get("current_assets", {})
                cash_bank = current_assets.get("13. Cash and bank balances", {})
                for key, value in cash_bank.items():
                    if key != "_metadata" and value is not None:
                        v24, v23 = self.get_value_flexible(value)
                        val_2024 += v24
                        val_2023 += v23
            
            elif item_name == "Short-term loans and advances":
                st_loans = company_data.get("loans_and_advances", {}).get("14. Short Term Loans and Advances", {})
                for key, value in st_loans.items():
                    if key != "_metadata" and value is not None:
                        v24, v23 = self.get_value_flexible(value)
                        val_2024 += v24
                        val_2023 += v23
            
            elif item_name == "Other current assets":
                other_ca = company_data.get("other_data", {}).get("15. Other Current Assets", {})
                for key, value in other_ca.items():
                    if key != "_metadata" and value is not None:
                        v24, v23 = self.get_value_flexible(value)
                        val_2024 += v24
                        val_2023 += v23
            
            # Add item if it has any value or is part of template structure
            if val_2024 != 0 or val_2023 != 0 or True:  # Always add template items
                items.append({
                    "category": category,
                    "subcategory": subcategory,
                    "name": item_name,
                    "note": note,
                    "value_2024": val_2024,
                    "value_2023": val_2023
                })
        
        return items

    def calculate_totals(self, items: List[Dict[str, Any]]) -> BalanceSheetTotals:
        """Calculate section totals and verify balance using template categories"""
        # Group by categories
        categories = {}
        for item in items:
            cat = item["category"]
            if cat not in categories:
                categories[cat] = {"2024": 0, "2023": 0}
            categories[cat]["2024"] += item["value_2024"]
            categories[cat]["2023"] += item["value_2023"]
        
        # Calculate major totals using template categories
        shareholders_funds_2024 = categories.get("Shareholders' funds", {}).get("2024", 0)
        shareholders_funds_2023 = categories.get("Shareholders' funds", {}).get("2023", 0)
        
        share_app_money_2024 = categories.get("Share application money pending allotment", {}).get("2024", 0)
        share_app_money_2023 = categories.get("Share application money pending allotment", {}).get("2023", 0)
        
        non_current_liab_2024 = categories.get("Non-Current liabilities", {}).get("2024", 0)
        non_current_liab_2023 = categories.get("Non-Current liabilities", {}).get("2023", 0)
        
        current_liab_2024 = categories.get("Current liabilities", {}).get("2024", 0)
        current_liab_2023 = categories.get("Current liabilities", {}).get("2023", 0)
        
        non_current_assets_2024 = categories.get("Non-current assets", {}).get("2024", 0)
        non_current_assets_2023 = categories.get("Non-current assets", {}).get("2023", 0)
        
        current_assets_2024 = categories.get("Current assets", {}).get("2024", 0)
        current_assets_2023 = categories.get("Current assets", {}).get("2023", 0)
        
        total_equity_liab_2024 = shareholders_funds_2024 + share_app_money_2024 + non_current_liab_2024 + current_liab_2024
        total_equity_liab_2023 = shareholders_funds_2023 + share_app_money_2023 + non_current_liab_2023 + current_liab_2023
        
        total_assets_2024 = non_current_assets_2024 + current_assets_2024
        total_assets_2023 = non_current_assets_2023 + current_assets_2023
        
        return BalanceSheetTotals(
            shareholders_funds_2024=shareholders_funds_2024,
            shareholders_funds_2023=shareholders_funds_2023,
            share_application_money_2024=share_app_money_2024,
            share_application_money_2023=share_app_money_2023,
            non_current_liabilities_2024=non_current_liab_2024,
            non_current_liabilities_2023=non_current_liab_2023,
            current_liabilities_2024=current_liab_2024,
            current_liabilities_2023=current_liab_2023,
            non_current_assets_2024=non_current_assets_2024,
            non_current_assets_2023=non_current_assets_2023,
            current_assets_2024=current_assets_2024,
            current_assets_2023=current_assets_2023,
            total_equity_liabilities_2024=total_equity_liab_2024,
            total_equity_liabilities_2023=total_equity_liab_2023,
            total_assets_2024=total_assets_2024,
            total_assets_2023=total_assets_2023,
            balance_difference_2024=abs(total_assets_2024 - total_equity_liab_2024),
            balance_difference_2023=abs(total_assets_2023 - total_equity_liab_2023)
        )

    def generate_balance_sheet_excel(self, items: List[Dict[str, Any]], totals: BalanceSheetTotals, output_dir: str = "output") -> str:
        """Generate formatted Excel balance sheet using template formatting"""
        os.makedirs(output_dir, exist_ok=True)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Balance Sheet"
        
        # Set column widths
        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        
        # Styles
        bold_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        row = 1
        
        def add_row(desc, note, val_2024, val_2023, bold=False, indent=0, border=False):
            nonlocal row
            
            # Description
            cell_a = ws.cell(row=row, column=1, value="  " * indent + desc)
            if bold:
                cell_a.font = bold_font
            if border:
                cell_a.border = thin_border
            
            # Note
            cell_b = ws.cell(row=row, column=2, value=note)
            if bold:
                cell_b.font = bold_font
            if border:
                cell_b.border = thin_border
            
            # Values
            for col, val in [(3, val_2024), (4, val_2023)]:
                cell = ws.cell(row=row, column=col)
                if val != 0:
                    cell.value = val
                    cell.number_format = '#,##0.00'
                if bold:
                    cell.font = bold_font
                if border:
                    cell.border = thin_border
                cell.alignment = Alignment(horizontal='right')
            
            row += 1
        
        # Header using template formatting
        header = self.formatting_rules.header
        add_row(header["title"], "", 0, 0, True)
        add_row("", "", 0, 0)
        add_row(header["currency_note"], "", 0, 0)
        add_row("", "", 0, 0)
        
        # Column headers
        headers = header["column_headers"]
        add_row(headers[0], headers[1], headers[2], headers[3], True)
        add_row("", "", 0, 0)
        
        # I. EQUITY AND LIABILITIES
        add_row("I. EQUITY AND LIABILITIES", "", 0, 0, True)
        add_row("", "", 0, 0)
        
        # (1) Shareholders' funds
        add_row("(1) Shareholders' funds", "", 0, 0, True)
        shareholders_items = [item for item in items if item["category"] == "Shareholders' funds"]
        for item in shareholders_items:
            add_row(f"    ({item['note'][0] if item['note'] else ''}) {item['name']}", item["note"], item["value_2024"], item["value_2023"])
        add_row("", "", totals.shareholders_funds_2024, totals.shareholders_funds_2023, True)
        add_row("", "", 0, 0)
        
        # (2) Share application money pending allotment
        share_app_items = [item for item in items if item["category"] == "Share application money pending allotment"]
        if any(item["value_2024"] != 0 or item["value_2023"] != 0 for item in share_app_items):
            add_row("(2) Share application money pending allotment", "", 0, 0, True)
            for item in share_app_items:
                add_row(f"    ({item['note']}) {item['name']}", item["note"], item["value_2024"], item["value_2023"])
            add_row("", "", totals.share_application_money_2024, totals.share_application_money_2023, True)
            add_row("", "", 0, 0)
        
        # (3) Non-Current liabilities
        add_row("(3) Non-Current liabilities", "", 0, 0, True)
        non_current_liab_items = [item for item in items if item["category"] == "Non-Current liabilities"]
        for item in non_current_liab_items:
            add_row(f"    ({item['note']}) {item['name']}", item["note"], item["value_2024"], item["value_2023"])
        add_row("", "", totals.non_current_liabilities_2024, totals.non_current_liabilities_2023, True)
        add_row("", "", 0, 0)
        
        # (4) Current liabilities
        add_row("(4) Current liabilities", "", 0, 0, True)
        current_liab_items = [item for item in items if item["category"] == "Current liabilities"]
        
        # Group trade payables
        trade_payables_items = [item for item in current_liab_items if item["subcategory"] == "Trade payables"]
        other_current_items = [item for item in current_liab_items if item["subcategory"] != "Trade payables"]
        
        # Add other current liability items first
        for item in other_current_items:
            add_row(f"    ({item['note']}) {item['name']}", item["note"], item["value_2024"], item["value_2023"])
        
        # Add trade payables with subcategory
        if trade_payables_items:
            trade_payables_total_2024 = sum(item["value_2024"] for item in trade_payables_items)
            trade_payables_total_2023 = sum(item["value_2023"] for item in trade_payables_items)
            
            add_row("    (11) Trade payables", "11", 0, 0, True, 1)
            for item in trade_payables_items:
                add_row(f"        (A) {item['name']}", item["note"], item["value_2024"], item["value_2023"])
            add_row("", "", trade_payables_total_2024, trade_payables_total_2023, True, 2)
        
        add_row("", "", totals.current_liabilities_2024, totals.current_liabilities_2023, True)
        add_row("", "", 0, 0)
        
        # TOTAL EQUITY & LIABILITIES
        add_row("TOTAL", "", totals.total_equity_liabilities_2024, totals.total_equity_liabilities_2023, True, 0, True)
        add_row("", "", 0, 0)
        
        # II. ASSETS
        add_row("II. ASSETS", "", 0, 0, True)
        add_row("", "", 0, 0)
        
        # Non-current assets
        add_row("Non-current assets", "", 0, 0, True)
        
        # (1) Property, Plant and Equipment
        ppe_items = [item for item in items if item.get("subcategory") == "Property, Plant and Equipment"]
        if ppe_items:
            add_row("(1) Property, Plant and Equipment", "", 0, 0, True, 1)
            ppe_total_2024 = ppe_total_2023 = 0
            for item in ppe_items:
                add_row(f"    ({item['note']}) {item['name']}", item["note"], item["value_2024"], item["value_2023"], False, 1)
                ppe_total_2024 += item["value_2024"]
                ppe_total_2023 += item["value_2023"]
            add_row("", "", ppe_total_2024, ppe_total_2023, True, 1)
            add_row("", "", 0, 0)
        
        # Other non-current assets
        other_non_current = [item for item in items if item["category"] == "Non-current assets" and item.get("subcategory") != "Property, Plant and Equipment"]
        for item in other_non_current:
            add_row(f"({item['note']}) {item['name']}", item["note"], item["value_2024"], item["value_2023"], False, 1)
        
        add_row("", "", totals.non_current_assets_2024, totals.non_current_assets_2023, True)
        add_row("", "", 0, 0)
        
        # (2) Current assets
        add_row("(2) Current assets", "", 0, 0, True)
        current_asset_items = [item for item in items if item["category"] == "Current assets"]
        for item in current_asset_items:
            add_row(f"    ({item['note']}) {item['name']}", item["note"], item["value_2024"], item["value_2023"], False, 1)
        
        add_row("", "", totals.current_assets_2024, totals.current_assets_2023, True)
        add_row("", "", 0, 0)
        
        # TOTAL ASSETS
        add_row("TOTAL", "", totals.total_assets_2024, totals.total_assets_2023, True, 0, True)
        
        # Add balance verification
        add_row("", "", 0, 0)
        balance_2024 = totals.balance_difference_2024
        balance_2023 = totals.balance_difference_2023
        
        if balance_2024 < 0.01 and balance_2023 < 0.01:
            add_row("✓ Balance Sheet is BALANCED", "", 0, 0, True)
        else:
            add_row(f"⚠ Balance Difference: {balance_2024:.2f} | {balance_2023:.2f}", "", 0, 0, True)
        
        # Save file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = os.path.join(output_dir, f"balance_sheet_{timestamp}.xlsx")
        wb.save(output_file)
        logger.info(f"Output file: {output_file}")
        print(f"Output file: {os.path.abspath(output_file)}")  # For API subprocess parsing
        return output_file

    def process(self, input_file: str, output_dir: str = "output") -> Optional[str]:
        """Main processing function"""
        try:
            logger.info(f"Processing: {input_file}")
            
            # Load JSON data
            with open(input_file, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
            
            logger.info("Extracting data using template structure...")
            
            # Method 1: Direct extraction using template structure
            items = self.extract_from_json_structure(json_data)
            logger.info(f"Extracted {len(items)} items using template structure")
            
            # Method 2: AI-assisted extraction if needed
            if len([item for item in items if item["value_2024"] != 0 or item["value_2023"] != 0]) < 5:
                logger.info("Using AI for additional extraction...")
                
                # Create summary for AI
                summary = json.dumps(json_data, indent=2)[:8000]  # Limit size
                ai_result = self.call_ai_for_analysis(summary)
                
                ai_items = ai_result.get("balance_sheet_items", [])
                logger.info(f"AI extracted {len(ai_items)} additional items")
                
                # Merge items (replace template items with AI items if they have values)
                ai_items_dict = {item["name"]: item for item in ai_items}
                
                for i, item in enumerate(items):
                    if item["name"] in ai_items_dict:
                        ai_item = ai_items_dict[item["name"]]
                        if ai_item["value_2024"] != 0 or ai_item["value_2023"] != 0:
                            items[i] = ai_item
            
            # Calculate totals
            totals = self.calculate_totals(items)
            
            # Display summary
            logger.info(f"\n BALANCE SHEET SUMMARY:")
            logger.info(f"Template Items: {len(self.template.get_template_structure())}")
            logger.info(f"Items with Values: {len([item for item in items if item['value_2024'] != 0 or item['value_2023'] != 0])}")
            logger.info(f" EQUITY & LIABILITIES 2024:")
            logger.info(f"  - Shareholders' funds: Rs. {totals.shareholders_funds_2024:,.2f} Lakhs")
            logger.info(f"  - Share application money: Rs. {totals.share_application_money_2024:,.2f} Lakhs")
            logger.info(f"  - Non-current liabilities: Rs. {totals.non_current_liabilities_2024:,.2f} Lakhs")
            logger.info(f"  - Current liabilities: Rs. {totals.current_liabilities_2024:,.2f} Lakhs")
            logger.info(f"  - TOTAL: Rs. {totals.total_equity_liabilities_2024:,.2f} Lakhs")
            logger.info(f" ASSETS 2024:")
            logger.info(f"  - Non-current assets: Rs. {totals.non_current_assets_2024:,.2f} Lakhs")
            logger.info(f"  - Current assets: Rs. {totals.current_assets_2024:,.2f} Lakhs")
            logger.info(f"  - TOTAL: Rs. {totals.total_assets_2024:,.2f} Lakhs")
            logger.info(f" Balance Difference 2024: Rs. {totals.balance_difference_2024:,.2f} Lakhs")
            logger.info(f" Balance Difference 2023: Rs. {totals.balance_difference_2023:,.2f} Lakhs")

            # Check if balanced
            is_balanced_2024 = totals.balance_difference_2024 < 0.01
            is_balanced_2023 = totals.balance_difference_2023 < 0.01
            
            if is_balanced_2024 and is_balanced_2023:
                logger.info(" Balance Sheet is PERFECTLY BALANCED!")
            else:
                logger.warning(" Balance Sheet has differences - may need adjustment")
            
            # Generate Excel using template formatting
            output_file = self.generate_balance_sheet_excel(items, totals, output_dir)
            
            logger.info(f" SUCCESS: Generated {output_file}")
            return output_file
            
        except Exception as e:
            logger.error(f" Error processing file: {e}", exc_info=True)
            return None

def main() -> None:
    """
    Main function for running the balance sheet generator.
    Accepts input file and output directory from command-line arguments or environment variables.
    Handles errors gracefully and logs all major events.
    """
    logger.info(" ENHANCED BALANCE SHEET GENERATOR v3.0 (Template-Based) started.")
    import sys
    api_key = settings.api_key
    input_file = settings.input_file
    output_dir = settings.output_dir
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
    if len(sys.argv) > 2:
        output_dir = sys.argv[2]
    if not api_key:
        logger.error(" Missing OPENROUTER_API_KEY environment variable. Please set your OpenRouter API key in the .env file.")
        return
    if not os.path.exists(input_file):
        logger.error(f" Input file not found: {input_file}. Please ensure your JSON data file exists.")
        return
    
    generator = EnhancedBalanceSheetGenerator(api_key)
    try:
        result = generator.process(input_file, output_dir)
        if result:
            abs_path = os.path.abspath(result)
            if os.path.exists(abs_path):
                logger.info(f" COMPLETED SUCCESSFULLY! Output file: {abs_path}")
                print(f"Output file: {abs_path}")  # For API subprocess parsing
            else:
                logger.error(f" PROCESSING FAILED. Output file not created: {abs_path}")
        else:
            logger.error(" PROCESSING FAILED. Please check the error messages above and try again.")
    except Exception as e:
        logger.error(f" Fatal error: {e}", exc_info=True)

if __name__ == "__main__":
    main()