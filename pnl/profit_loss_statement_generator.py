import os
import json
import logging
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from typing import Dict, List, Tuple, Any, Optional
from pydantic import BaseModel, Field, ValidationError
from pydantic_settings import BaseSettings

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class Settings(BaseSettings):
    """Settings for P&L generation, loaded from environment variables or .env file."""
    json_files: List[str] = Field(default_factory=lambda: [
        "clean_financial_data_pnl.json",
        "pnl_notes.json"
    ], env="PNL_JSON_FILES")
    output_file: str = Field(default="data/pnl_statement.xlsx", env="PNL_OUTPUT_FILE")

settings = Settings()

class FinancialItem(BaseModel):
    name: str
    values: List[float] = Field(default_factory=list)

class FinancialDataModel(BaseModel):
    other_data: Dict[str, Any] = Field(default_factory=dict)

class PnLGenerator:
    def __init__(self, json_file_path: str = settings.json_files[0]):
        """Initialize the P&L generator with JSON file path."""
        self.json_file_path = json_file_path
        self.financial_data: Dict[str, Any] = {}

    def load_financial_data(self) -> bool:
        """Load financial data from JSON file."""
        try:
            logger.info(f"Loading financial data from: {self.json_file_path}")
            with open(self.json_file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            # Handle different JSON structures flexibly
            if "company_financial_data" in data:
                self.financial_data = data["company_financial_data"].get("other_data", {})
            elif "other_data" in data:
                self.financial_data = data["other_data"]
            else:
                self.financial_data = data
            logger.info(f"Loaded data for {len(self.financial_data)} financial items")
            return True
        except FileNotFoundError:
            logger.error(f"File not found: {self.json_file_path}")
            return False
        except json.JSONDecodeError as e:
            logger.error(f"Invalid JSON format: {str(e)}")
            return False
        except Exception as e:
            logger.error(f"Error loading data: {str(e)}")
            return False

    def extract_values(self, item_key: str) -> Tuple[float, float]:
        """Extract 2024 and 2023 values from financial data."""
        if item_key not in self.financial_data:
            logger.warning(f"{item_key} not found in data")
            return 0.0, 0.0
        item_data = self.financial_data[item_key]
        def recursive_sum(data):
            sum_2024, sum_2023 = 0.0, 0.0
            if isinstance(data, dict):
                for k, v in data.items():
                    # Skip metadata
                    if k == "_metadata":
                        continue
                    s24, s23 = recursive_sum(v)
                    sum_2024 += s24
                    sum_2023 += s23
            elif isinstance(data, list):
                # If list contains only numbers, try to use first two as 2024/2023
                nums = [x for x in data if isinstance(x, (int, float))]
                if len(nums) >= 2:
                    sum_2024 += float(nums[0] or 0)
                    sum_2023 += float(nums[1] or 0)
                elif len(nums) == 1:
                    sum_2024 += float(nums[0] or 0)
                # Otherwise, skip non-numeric entries
            elif isinstance(data, (int, float)):
                sum_2024 += float(data)
            elif isinstance(data, str):
                # Try to parse as float
                try:
                    val = float(data)
                    sum_2024 += val
                except Exception:
                    pass
            return sum_2024, sum_2023

        # Special handling for date-based dicts
        def sum_dates(data):
            sum_2024, sum_2023 = 0.0, 0.0
            if isinstance(data, dict):
                for k, v in data.items():
                    if k == "_metadata":
                        continue
                    if isinstance(v, dict):
                        # If keys look like dates, sum by year
                        for date_key, val in v.items():
                            if "2024" in date_key:
                                try:
                                    sum_2024 += float(val)
                                except Exception:
                                    pass
                            elif "2023" in date_key:
                                try:
                                    sum_2023 += float(val)
                                except Exception:
                                    pass
                    else:
                        s24, s23 = recursive_sum(v)
                        sum_2024 += s24
                        sum_2023 += s23
            return sum_2024, sum_2023

        # Try date-based sum first, fallback to recursive
        s24, s23 = sum_dates(item_data)
        if s24 == 0.0 and s23 == 0.0:
            s24, s23 = recursive_sum(item_data)
        logger.info(f"Extracted for {item_key}: 2024={s24}, 2023={s23}")
        return s24, s23

    def get_revenue_data(self) -> Tuple[float, float]:
        """Extract revenue from operations data."""
        return self.extract_values("16. Revenue from Operations")

    def get_other_income_data(self) -> Tuple[float, float]:
        """Extract other income data."""
        return self.extract_values("17. Other income")

    def get_cost_materials_data(self) -> Tuple[float, float]:
        """Extract cost of materials consumed data."""
        item_key = "18. Cost of materials consumed"
        if item_key not in self.financial_data:
            logger.warning(f"{item_key} not found in data")
            return 0.0, 0.0
        item_data = self.financial_data[item_key]
        if "Cost of materials consumed" in item_data:
            values = item_data["Cost of materials consumed"]
            if isinstance(values, list) and len(values) >= 2:
                return float(values[0] or 0), float(values[1] or 0)
        # Fallback: calculate from opening stock + purchases - closing stock
        opening_2024 = opening_2023 = 0.0
        purchases_2024 = purchases_2023 = 0.0
        closing_2024 = closing_2023 = 0.0
        if "Opening stock" in item_data:
            values = item_data["Opening stock"]
            if isinstance(values, list) and len(values) >= 2:
                opening_2024, opening_2023 = float(values[0] or 0), float(values[1] or 0)
        if "Add: Purchases" in item_data:
            values = item_data["Add: Purchases"]
            if isinstance(values, list) and len(values) >= 2:
                purchases_2024, purchases_2023 = float(values[0] or 0), float(values[1] or 0)
        if "Less: Closing stock" in item_data:
            values = item_data["Less: Closing stock"]
            if isinstance(values, list) and len(values) >= 2:
                closing_2024, closing_2023 = float(values[0] or 0), float(values[1] or 0)
        cost_2024 = opening_2024 + purchases_2024 - closing_2024
        cost_2023 = opening_2023 + purchases_2023 - closing_2023
        return cost_2024, cost_2023

    def get_employee_expense_data(self) -> Tuple[float, float]:
        """Extract employee benefit expense data."""
        return self.extract_values("19. Employee benefit expense")

    def get_other_expenses_data(self) -> Tuple[float, float]:
        """Extract other expenses data."""
        return self.extract_values("20. Other expenses")

    def get_depreciation_data(self) -> Tuple[float, float]:
        """Extract depreciation and amortisation data."""
        return self.extract_values("21. Depreciation and amortisation expense")

    def get_loss_on_sale_data(self) -> Tuple[float, float]:
        """Extract loss on sale of assets data."""
        return self.extract_values("22. Loss on sale of assets")

    def get_finance_costs_data(self) -> Tuple[float, float]:
        """Extract finance costs data."""
        return self.extract_values("23. Finance costs")

    def format_currency(self, value: float) -> str:
        """Format currency with commas."""
        if value == 0:
            return ""
        return f"{value:,.2f}"

    def generate_pnl_statement(self, output_file: str = settings.output_file) -> bool:
        """Generate comprehensive P&L statement Excel file."""
        if not self.financial_data:
            logger.error("No financial data loaded. Please load data first.")
            return False
        wb = Workbook()
        ws = wb.active
        ws.title = "Profit and Loss Statement"
        title_font = Font(bold=True, size=12)
        header_font = Font(bold=True, size=10)
        normal_font = Font(size=10)
        bold_font = Font(bold=True, size=10)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        top_bottom_border = Border(
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        row = 1
        ws.merge_cells("A1:D1")
        ws["A1"] = "PART II - STATEMENT OF PROFIT AND LOSS"
        ws["A1"].font = title_font
        ws["A1"].alignment = center_align
        row += 2
        ws["A3"] = "Name of the Company............................"
        ws["A3"].font = normal_font
        ws["A3"].alignment = left_align
        row += 1
        ws["A4"] = "Profit and loss statement for the year ended ..........................."
        ws["A4"].font = normal_font
        ws["A4"].alignment = left_align
        row += 2
        ws["D6"] = "(Rupees in...........)"
        ws["D6"].font = normal_font
        ws["D6"].alignment = right_align
        row += 2

        # Table headers
        headers = ["Particulars", "Note No.", "Figures as at the end of current reporting period", "Figures as at the end of the previous reporting period"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = center_align if col > 1 else left_align
        row += 1
        
        # Column numbers
        col_numbers = ["1", "2", "3", "4"]
        for col, num in enumerate(col_numbers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = num
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = center_align
        row += 1

        def add_data_row(description: str, note_ref: str, val_2024: float, val_2023: float,
                         is_bold: bool = False) -> None:
            """Add a data row with proper formatting."""
            nonlocal row
            cell_a = ws.cell(row=row, column=1)
            cell_a.value = description
            cell_a.font = bold_font if is_bold else normal_font
            cell_a.alignment = left_align
            cell_a.border = thin_border

            cell_b = ws.cell(row=row, column=2)
            cell_b.value = note_ref if note_ref else ""
            cell_b.font = normal_font
            cell_b.alignment = center_align
            cell_b.border = thin_border

            cell_c = ws.cell(row=row, column=3)
            cell_c.value = "xxx" if val_2024 == 0 and description != "Total Revenue (I + II)" and description != "Total expenses" and "Total" not in description and "Profit" not in description else self.format_currency(val_2024)
            cell_c.font = bold_font if is_bold else normal_font
            cell_c.alignment = center_align
            cell_c.border = thin_border

            cell_d = ws.cell(row=row, column=4)
            cell_d.value = "xxx" if val_2023 == 0 and description != "Total Revenue (I + II)" and description != "Total expenses" and "Total" not in description and "Profit" not in description else self.format_currency(val_2023)
            cell_d.font = bold_font if is_bold else normal_font
            cell_d.alignment = center_align
            cell_d.border = thin_border
            row += 1

        logger.info("Extracting financial data...")
        revenue_2024, revenue_2023 = self.get_revenue_data()
        other_income_2024, other_income_2023 = self.get_other_income_data()
        materials_2024, materials_2023 = self.get_cost_materials_data()
        employee_2024, employee_2023 = self.get_employee_expense_data()
        other_exp_2024, other_exp_2023 = self.get_other_expenses_data()
        depreciation_2024, depreciation_2023 = self.get_depreciation_data()
        loss_sale_2024, loss_sale_2023 = self.get_loss_on_sale_data()
        finance_2024, finance_2023 = self.get_finance_costs_data()

        # Revenue Section
        add_data_row("I. Revenue from operations", "16", revenue_2024, revenue_2023)
        add_data_row("II. Other income", "17", other_income_2024, other_income_2023)
        total_revenue_2024 = revenue_2024 + other_income_2024
        total_revenue_2023 = revenue_2023 + other_income_2023
        add_data_row("III. Total Revenue (I + II)", "", total_revenue_2024, total_revenue_2023, is_bold=True)

        # Expenses Section
        add_data_row("IV. Expenses:", "", 0, 0)
        add_data_row("Cost of materials consumed", "18", materials_2024, materials_2023)
        add_data_row("Purchases of Stock-in-Trade", "", 0, 0)
        add_data_row("Changes in inventories of finished goods", "", 0, 0)
        add_data_row("work-in-progress and", "", 0, 0)
        add_data_row("Stock-in-Trade", "", 0, 0)
        add_data_row("Employee benefits expense", "19", employee_2024, employee_2023)
        add_data_row("Finance costs", "23", finance_2024, finance_2023)
        add_data_row("Depreciation and amortisation expense", "21", depreciation_2024, depreciation_2023)
        add_data_row("Other expenses", "20", other_exp_2024, other_exp_2023)

        total_expenses_2024 = materials_2024 + employee_2024 + other_exp_2024 + depreciation_2024 + loss_sale_2024 + finance_2024
        total_expenses_2023 = materials_2023 + employee_2023 + other_exp_2023 + depreciation_2023 + loss_sale_2023 + finance_2023
        add_data_row("Total expenses", "", total_expenses_2024, total_expenses_2023, is_bold=True)

        profit_before_exceptional_2024 = total_revenue_2024 - total_expenses_2024
        profit_before_exceptional_2023 = total_revenue_2023 - total_expenses_2023
        add_data_row("V. Profit before exceptional and extraordinary items and tax (III - IV)", "", profit_before_exceptional_2024, profit_before_exceptional_2023, is_bold=True)

        add_data_row("VI. Exceptional items", "", 0, 0)
        add_data_row("VII. Profit before extraordinary items and tax (V - VI)", "", profit_before_exceptional_2024, profit_before_exceptional_2023, is_bold=True)
        add_data_row("VIII. Extraordinary items", "", 0, 0)
        add_data_row("IX. Profit before tax (VII- VIII)", "", profit_before_exceptional_2024, profit_before_exceptional_2023, is_bold=True)

        # Tax Section
        add_data_row("X. Tax expense:", "", 0, 0)
        add_data_row("(1) Current tax", "", 0, 0)
        add_data_row("(2) Deferred tax", "", 0, 0)

        add_data_row("XI. Profit (Loss) for the period from continuing operations (VII-VIII)", "", profit_before_exceptional_2024, profit_before_exceptional_2023, is_bold=True)
        add_data_row("XII. Profit/(loss) from discontinuing operations", "", 0, 0)
        add_data_row("XIII. Tax expense of discontinuing operations", "", 0, 0)
        add_data_row("XIV. Profit/(loss) from Discontinuing operations (after tax) (XII-XIII)", "", 0, 0)
        add_data_row("XV. Profit (Loss) for the period (XI + XIV)", "", profit_before_exceptional_2024, profit_before_exceptional_2023, is_bold=True)

        # Earnings per share
        add_data_row("XVI. Earnings per equity share:", "", 0, 0)
        add_data_row("(1) Basic", "", 0, 0)
        add_data_row("(2) Diluted", "", 0, 0)

        # Footer
        row += 2
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"] = "See accompanying notes to the financial statements."
        ws[f"A{row}"].font = normal_font
        ws[f"A{row}"].alignment = left_align
        row += 1
        
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"] = "GENERAL INSTRUCTIONS FOR PREPARATION OF STATEMENT OF"
        ws[f"A{row}"].font = title_font
        ws[f"A{row}"].alignment = center_align

        # Save the file
        try:
            wb.save(output_file)
            logger.info(f"P&L Statement generated successfully: {output_file}")
            print(f"Output file: {os.path.abspath(output_file)}")  # For API subprocess parsing
            self.print_financial_summary(
                total_revenue_2024, total_revenue_2023,
                total_expenses_2024, total_expenses_2023,
                profit_before_exceptional_2024, profit_before_exceptional_2023,
                profit_before_exceptional_2024, profit_before_exceptional_2023
            )
            return True
        except PermissionError:
            logger.error(f"Permission Error: Cannot save to {output_file}")
            fallback_file = os.path.join(os.path.expanduser("~"), "Desktop", "pnl_statement_fallback.xlsx")
            try:
                wb.save(fallback_file)
                logger.info(f"P&L Statement saved to: {fallback_file}")
                print(f"Output file: {os.path.abspath(fallback_file)}")  # For API subprocess parsing
                return True
            except Exception as e:
                logger.error(f"Failed to save: {str(e)}")
                return False
        except Exception as e:
            logger.error(f"Error saving file: {str(e)}")
            return False

    def print_financial_summary(self, total_revenue_2024: float, total_revenue_2023: float,
                                total_expenses_2024: float, total_expenses_2023: float,
                                profit_before_tax_2024: float, profit_before_tax_2023: float,
                                profit_after_tax_2024: float, profit_after_tax_2023: float) -> None:
        """Log financial summary."""
        logger.info("=" * 60)
        logger.info("FINANCIAL SUMMARY")
        logger.info("=" * 60)
        logger.info(f"Total Revenue 2024:     Rs.{total_revenue_2024:>12,.2f} Lakhs")
        logger.info(f"Total Revenue 2023:     Rs.{total_revenue_2023:>12,.2f} Lakhs")
        logger.info(f"Total Expenses 2024:    Rs.{total_expenses_2024:>12,.2f} Lakhs")
        logger.info(f"Total Expenses 2023:    Rs.{total_expenses_2023:>12,.2f} Lakhs")
        logger.info(f"Profit Before Tax 2024: Rs.{profit_before_tax_2024:>12,.2f} Lakhs")
        logger.info(f"Profit Before Tax 2023: Rs.{profit_before_tax_2023:>12,.2f} Lakhs")
        logger.info(f"Profit After Tax 2024:  Rs.{profit_after_tax_2024:>12,.2f} Lakhs")
        logger.info(f"Profit After Tax 2023:  Rs.{profit_after_tax_2023:>12,.2f} Lakhs")
        if total_revenue_2023 > 0:
            growth_rate = ((total_revenue_2024 - total_revenue_2023) / total_revenue_2023) * 100
            logger.info(f"Revenue Growth Rate:    {growth_rate:>12.2f}%")

def main() -> None:
    logger.info("P&L STATEMENT GENERATOR FROM JSON")
    logger.info("=" * 50)
    logger.info(f"Current working directory: {os.getcwd()}")

    # Determine input JSON file (env, arg, or default)
    json_file = os.getenv("PNL_INPUT_FILE", None)
    if not json_file:
        if len(sys.argv) > 1:
            json_file = sys.argv[1]
            logger.info(f"Input JSON file from argument: {json_file}")
        else:
            for file in settings.json_files:
                if os.path.exists(file):
                    json_file = file
                    logger.info(f"Found input JSON file: {json_file}")
                    break
    if not json_file or not os.path.exists(json_file):
        logger.error(f"Input JSON file '{json_file}' not found. Please provide a valid file.")
        return

    # Determine output Excel file (env, arg, or default)
    output_path = os.getenv("PNL_OUTPUT_FILE", settings.output_file)
    if len(sys.argv) > 2:
        output_path = sys.argv[2]
        logger.info(f"Output Excel path from argument: {output_path}")
    logger.info(f"Output file: {output_path}")

    generator = PnLGenerator(json_file)
    if generator.load_financial_data():
        try:
            if generator.generate_pnl_statement(output_path):
                logger.info(f"P&L Statement generated successfully: {os.path.abspath(output_path)}")
            else:
                logger.error("Failed to generate P&L statement.")
        except Exception as e:
            logger.error(f"Error writing Excel file: {e}")
    else:
        logger.error("Failed to load financial data")

if __name__ == "__main__":
    main()