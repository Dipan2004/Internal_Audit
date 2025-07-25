import os
import json
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
import re
import uuid

INPUT_FILE = "data/BLnotes2.xlsx"
OUTPUT_FILE = "BL_Sheet2.xlsx"

def safe_float_conversion(value):
    """Safely convert value to float, handling various formats including #REF! errors."""
    if value is None:
        return 0.0
    
    str_val = str(value).strip()
    
    # Handle Excel errors and empty values
    if not str_val or str_val in ['-', '--', '', 'None', '#REF!', '#DIV/0!', '#VALUE!', '#NAME?', '#N/A', 'NA', 'nil', 'Nil', 'NIL']:
        return 0.0
    
    # Remove common formatting
    str_val = str_val.replace(',', '').replace('₹', '').replace('Rs.', '').replace(' ', '')
    str_val = str_val.replace('INR', '').replace('Lakhs', '').replace('lakhs', '').replace('LAKHS', '')
    
    # Handle parentheses as negative
    is_negative = '(' in str_val and ')' in str_val
    str_val = str_val.replace('(', '').replace(')', '')
    
    # Remove any remaining non-numeric characters except decimal point and minus
    str_val = re.sub(r'[^\d.-]', '', str_val)
    
    if not str_val or str_val == '-' or str_val == '.':
        return 0.0
    
    try:
        result = float(str_val)
        return -result if is_negative else result
    except (ValueError, TypeError):
        return 0.0

def find_data_columns(ws):
    """Find the data columns for 2024 and 2023 with enhanced detection."""
    col_2024 = None
    col_2023 = None
    
    print("🔍 Searching for year columns...")
    
    # Look for year headers in first 30 rows and 20 columns
    for row_idx in range(1, min(ws.max_row + 1, 30)):
        for col_idx in range(1, min(ws.max_column + 1, 20)):
            cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "").strip()
            
            # Enhanced year detection patterns
            patterns_2024 = [
                r'(31\.?03\.?2024|march.*2024|2024.*march|45382|2024)',
                r'march\s*31,?\s*2024',
                r'2024\s*-?\s*2023',  # Sometimes both years in one cell
            ]
            
            patterns_2023 = [
                r'(31\.?03\.?2023|march.*2023|2023.*march|45016|2023)',
                r'march\s*31,?\s*2023',
            ]
            
            # Check for 2024
            for pattern in patterns_2024:
                if re.search(pattern, cell_value, re.IGNORECASE):
                    col_2024 = col_idx
                    print(f"  ✅ Found 2024 column at: {col_idx} (value: '{cell_value}')")
                    break
            
            # Check for 2023
            for pattern in patterns_2023:
                if re.search(pattern, cell_value, re.IGNORECASE):
                    col_2023 = col_idx
                    print(f"  ✅ Found 2023 column at: {col_idx} (value: '{cell_value}')")
                    break
    
    # Enhanced fallback detection by analyzing column content
    if col_2024 is None or col_2023 is None:
        print("  🔍 Year headers not found, analyzing column patterns...")
        
        # Analyze columns for numeric data patterns
        column_stats = {}
        for col_idx in range(2, min(ws.max_column + 1, 15)):  # Skip first column (labels)
            numeric_count = 0
            total_value = 0
            non_zero_count = 0
            ref_error_count = 0
            
            for row_idx in range(1, min(ws.max_row + 1, 100)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    str_val = str(cell_value).strip()
                    if '#REF!' in str_val or '#DIV/0!' in str_val:
                        ref_error_count += 1
                    converted = safe_float_conversion(cell_value)
                    if converted != 0.0:
                        numeric_count += 1
                        non_zero_count += 1
                        total_value += abs(converted)
                    elif str_val and str_val != '#REF!':
                        numeric_count += 1
            
            if numeric_count > 3:  # Column has meaningful data
                column_stats[col_idx] = {
                    'numeric_count': numeric_count,
                    'non_zero_count': non_zero_count,
                    'total_value': total_value,
                    'ref_error_count': ref_error_count,
                    'avg_value': total_value / non_zero_count if non_zero_count > 0 else 0
                }
                print(f"    Column {col_idx}: numeric={numeric_count}, non-zero={non_zero_count}, "
                      f"total={total_value:.2f}, ref_errors={ref_error_count}")
        
        # Smart column assignment
        sorted_cols = sorted(column_stats.keys())
        if len(sorted_cols) >= 2:
            # Prefer columns with fewer #REF! errors for current year (2024)
            best_2024_col = min(sorted_cols, key=lambda x: column_stats[x]['ref_error_count'])
            remaining_cols = [c for c in sorted_cols if c != best_2024_col]
            
            if col_2024 is None:
                col_2024 = best_2024_col
                print(f"  🎯 Assigned 2024 column: {col_2024} (ref_errors: {column_stats[best_2024_col]['ref_error_count']})")
            
            if col_2023 is None and remaining_cols:
                col_2023 = remaining_cols[0]
                print(f"  🎯 Assigned 2023 column: {col_2023}")
        
        # Ultimate fallback based on typical patterns
        if col_2024 is None and col_2023 is None:
            # Look for the most common pattern: data in columns C and D or B and C
            if len(sorted_cols) >= 2:
                col_2024 = sorted_cols[0]
                col_2023 = sorted_cols[1]
    
    # Default fallback
    if col_2024 is None:
        col_2024 = 3  # Column C (more common than D)
        print(f"  ⚠️ Using default 2024 column: {col_2024}")
    if col_2023 is None:
        col_2023 = 4  # Column D
        print(f"  ⚠️ Using default 2023 column: {col_2023}")
    
    print(f"  📊 Final columns - 2024: {col_2024}, 2023: {col_2023}")
    return col_2024, col_2023

def is_skip_row(text, row_values):
    """Enhanced row skipping logic with better pattern recognition."""
    if not text or len(text.strip()) <= 1:
        return True
    
    text_lower = text.lower().strip()
    
    # Enhanced skip patterns
    skip_patterns = [
        r'^(in\s+lakhs?|amount|particulars|year\s+ended|march\s+31|2024|2023)$',
        r'^(sr\.?\s*no\.?|s\.?\s*no\.?|note\s*no\.?|\#ref\!?)$',
        r'^\d+\s*-\s*\d+\s*(year|month|day)',  
        r'^(outstanding\s+for|age\s+wise|gross\s+carrying)',
        r'^(accumulated\s+depreciation|net\s+carrying)',
        r'^(terms/\s*rights|details\s+of|disclosure\s+of)',
        r'^(reconciliation|sundry|undisputed|disputed)',
        r'^\(?as\s+per\s+',  
        r'^\(?add\s*/\s*less\s*\)?',  
        r'^\(?previous\s+year\s*\)?',
        r'^(\(|\s)*\d+\s*(\)|\s)*$',  
        r'^(nil|na|n/a|not\s+applicable|none|zero)$',
        r'^(balance|opening|closing)$',
        r'^\s*-+\s*$',  # Lines with just dashes
        r'^\s*=+\s*$',  # Lines with just equals
        r'^particulars\s*$',
        r'^outstanding\s+for\s+following\s+periods',
    ]
    
    for pattern in skip_patterns:
        if re.match(pattern, text_lower):
            return True
    
    # Skip if text is mostly non-alphabetic
    alpha_count = sum(1 for c in text if c.isalpha())
    if len(text) > 3 and alpha_count < len(text) * 0.25:
        return True
    
    # Skip if all values in the row are zero/empty/errors
    if row_values and len(row_values) > 1:
        meaningful_values = []
        for val in row_values[1:]:  # Skip first column (label)
            if val is not None:
                str_val = str(val).strip()
                if str_val and str_val not in ['#REF!', '#DIV/0!', '#VALUE!', '0', '0.0', '-']:
                    converted = safe_float_conversion(val)
                    if converted != 0.0:
                        meaningful_values.append(converted)
        
        # If no meaningful values and label is not descriptive, skip
        if not meaningful_values and alpha_count < 5:
            return True
    
    return False

def extract_meaningful_data(text, value_2024, value_2023, note_num=None):
    """Enhanced data extraction with better label cleaning."""
    if not text:
        return None
    
    # Clean up the label
    cleaned_label = text.strip()
    
    # Remove common prefixes/suffixes
    cleanup_patterns = [
        r'^\d+\.\s*',  # Leading numbers like "1. "
        r'^\(?[a-z]\)\s*',  # Leading letters like "(a) "
        r'\s*\(.*lakhs?\).*$',  
        r'\s*-\s*in\s+lakhs?.*$',  
        r'^\d+\s+',  # Leading standalone numbers
        r'^\s*[-•]\s*',  # Leading bullets
    ]
    
    for pattern in cleanup_patterns:
        cleaned_label = re.sub(pattern, '', cleaned_label, flags=re.IGNORECASE).strip()
    
    # Skip if label becomes too short or meaningless
    if len(cleaned_label) < 2:
        return None
    
    # Enhanced meaningless entries check
    meaningless_entries = [
        'nil', 'na', 'n/a', 'not applicable', 'none', 'zero', '',
        'amount', 'value', 'sum',
        'march', 'year', 'period', 'date', 'particulars'
    ]
    
    if cleaned_label.lower() in meaningless_entries:
        return None
    
    # Special handling for specific notes
    special_cases = {
        "2": ["share capital", "equity shares", "authorised", "issued", "subscribed", "paid"],
        "3": ["reserves", "surplus", "securities premium", "profit", "loss"],
        "4": ["borrowings", "loan", "debt", "financial", "bank"],
        "9": ["assets", "building", "plant", "machinery", "equipment", "vehicle", "computer", "furniture"],
        "11": ["inventories", "stock", "consumables", "raw material"],
        "12": ["receivables", "debtors", "outstanding"],
        "13": ["cash", "bank", "deposit", "balance"],
    }
    
    # Check if this is a meaningful entry for the specific note
    has_values = value_2024 != 0.0 or value_2023 != 0.0
    has_meaningful_text = len(cleaned_label) > 3 and any(c.isalpha() for c in cleaned_label)
    
    # For specific notes, be more lenient with text-only entries
    is_relevant_for_note = False
    if note_num and str(note_num) in special_cases:
        relevant_keywords = special_cases[str(note_num)]
        is_relevant_for_note = any(keyword in cleaned_label.lower() for keyword in relevant_keywords)
    
    if has_values or (has_meaningful_text and (is_relevant_for_note or len(cleaned_label) > 8)):
        return {
            "label": cleaned_label,
            "value": value_2024,
            "previous_value": value_2023,
            "change": value_2024 - value_2023,
            "change_percent": ((value_2024 - value_2023) / value_2023 * 100) if value_2023 != 0 else 0
        }
    
    return None

def get_note_total_value(note_num, subcategories):
    """Get the actual total value from the note - look for total/balance lines or use highest value."""
    if not subcategories:
        return 0.0, 0.0
    
    # First, look for explicit total/balance lines
    total_indicators = [
        "total", "balance at the end", "closing balance", "net surplus",
        "grand total", "sum", "net carrying amount", "carrying amount"
    ]
    
    for sub in subcategories:
        label_lower = sub["label"].lower()
        if any(indicator in label_lower for indicator in total_indicators):
            print(f"  🎯 Found total line for Note {note_num}: '{sub['label']}' = 2024: {sub['value']:.2f}, 2023: {sub['previous_value']:.2f}")
            return sub["value"], sub["previous_value"]
    
    # If no explicit total found, look for the largest absolute value (likely the main item)
    if subcategories:
        max_entry = max(subcategories, key=lambda x: abs(x["value"]) + abs(x["previous_value"]))
        print(f"  📊 Using max value for Note {note_num}: '{max_entry['label']}' = 2024: {max_entry['value']:.2f}, 2023: {max_entry['previous_value']:.2f}")
        return max_entry["value"], max_entry["previous_value"]
    
    return 0.0, 0.0

def enhanced_note_detection(ws):
    """Enhanced note header detection with multiple strategies."""
    note_headers = {}
    
    print("🔍 Enhanced note detection starting...")
    
    # Strategy 1: Look for explicit "Note X" patterns
    for row_idx in range(1, min(ws.max_row + 1, 200)):
        first_cell = str(ws.cell(row=row_idx, column=1).value or "").strip()
        
        note_patterns = [
            r'^note\s*(\d{1,2})\s*[:\-\.]?\s*(.+)',  # "Note 1: Description"
            r'^(\d{1,2})\.\s*(.+)',                   # "1. Description"
            r'^(\d{1,2})\s*[:\-]\s*(.+)',            # "1: Description" or "1- Description"
            r'^(\d{1,2})\s+([a-zA-Z].{5,})',         # "1 Description" (at least 5 chars)
        ]
        
        for pattern in note_patterns:
            match = re.match(pattern, first_cell, re.IGNORECASE)
            if match:
                note_num = match.group(1)
                note_title = match.group(2).strip()
                
                # Validate note number and title
                if 1 <= int(note_num) <= 30 and len(note_title) > 3:
                    # Clean up title
                    note_title = re.sub(r'\s+', ' ', note_title)
                    note_headers[int(note_num)] = {
                        'row': row_idx,
                        'title': note_title,
                        'note_num': note_num
                    }
                    print(f"  📝 Found Note {note_num} at row {row_idx}: {note_title[:60]}")
                    break
    
    # Strategy 2: Look for common note titles even without explicit numbering
    common_note_titles = {
        "share capital": 2,
        "reserves and surplus": 3,
        "long term borrowings": 4,
        "long-term borrowings": 4,
        "deferred tax": 5,
        "trade payables": 6,
        "other current liabilities": 7,
        "short term provisions": 8,
        "short-term provisions": 8,
        "fixed assets": 9,
        "long term loans": 10,
        "inventories": 11,
        "trade receivables": 12,
        "cash and bank": 13,
        "short term loans": 14,
        "other current assets": 15,
    }
    
    # Look for these titles in the first few columns
    for row_idx in range(1, min(ws.max_row + 1, 200)):
        for col_idx in range(1, min(4, ws.max_column + 1)):  # Check first 3 columns
            cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "").strip().lower()
            
            for title_key, note_num in common_note_titles.items():
                if title_key in cell_value and note_num not in note_headers:
                    note_headers[note_num] = {
                        'row': row_idx,
                        'title': cell_value.title(),
                        'note_num': str(note_num)
                    }
                    print(f"  📝 Inferred Note {note_num} at row {row_idx}: {cell_value[:60]}")
                    break
    
    return note_headers

def load_and_map_excel_notes(file_path=INPUT_FILE):
    """Enhanced Excel loading with improved error handling and data extraction."""
    try:
        print(f"📖 Reading Excel file: {file_path}")
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        notes_data = {}
        
        print(f"📐 Sheet dimensions: {ws.max_row} rows × {ws.max_column} columns")
        
        col_2024, col_2023 = find_data_columns(ws)
        note_headers = enhanced_note_detection(ws)
        
        print(f"📋 Found {len(note_headers)} note headers")
        
        # Process each note
        for note_num in sorted(note_headers.keys()):
            note_info = note_headers[note_num]
            start_row = note_info['row']
            
            # Determine end row (next note or reasonable limit)
            end_row = start_row + 50  # Default range
            for next_note_num in sorted(note_headers.keys()):
                if next_note_num > note_num:
                    end_row = min(note_headers[next_note_num]['row'] - 1, start_row + 100)
                    break
            end_row = min(end_row, ws.max_row)
            
            print(f"\n📊 Processing Note {note_info['note_num']}: {note_info['title']}")
            print(f"   Rows {start_row} to {end_row}")
            
            subcategories = []
            
            # Extract data from this note's rows
            for row_idx in range(start_row + 1, end_row + 1):
                try:
                    # Get all row values
                    row_values = []
                    for col_idx in range(1, min(ws.max_column + 1, 20)):
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        row_values.append(cell_value)
                    
                    if not row_values or not row_values[0]:
                        continue
                    
                    first_cell = str(row_values[0]).strip()
                    
                    # Skip unwanted rows
                    if is_skip_row(first_cell, row_values):
                        continue
                    
                    # Extract values with bounds checking
                    value_2024 = 0.0
                    value_2023 = 0.0
                    
                    if col_2024 <= len(row_values):
                        value_2024 = safe_float_conversion(row_values[col_2024 - 1])
                    if col_2023 <= len(row_values):
                        value_2023 = safe_float_conversion(row_values[col_2023 - 1])
                    
                    # Extract meaningful data
                    data_entry = extract_meaningful_data(first_cell, value_2024, value_2023, note_info['note_num'])
                    if data_entry:
                        subcategories.append(data_entry)
                        print(f"    ✓ {data_entry['label'][:40]:<40} | 2024: {value_2024:>10.2f} | 2023: {value_2023:>10.2f}")
                
                except Exception as e:
                    print(f"    ❌ Error processing row {row_idx}: {str(e)}")
                    continue
            
            # Get the actual note total (not sum of all items)
            total_2024, total_2023 = get_note_total_value(note_info['note_num'], subcategories)
            
            # Store note data
            notes_data[note_info['note_num']] = {
                "category_name": note_info['title'],
                "structure": [{
                    "category": note_info['title'],
                    "subcategories": subcategories
                }],
                "total_2024": total_2024,
                "total_2023": total_2023,
                "total_change": total_2024 - total_2023
            }
            
            print(f"  📊 Note {note_info['note_num']} totals: 2024={total_2024:.2f}, 2023={total_2023:.2f}, Items={len(subcategories)}")
        
        # Save JSON
        json_file_path = "data/balance_sheet_notes.json"
        os.makedirs(os.path.dirname(json_file_path), exist_ok=True)
        with open(json_file_path, "w", encoding="utf-8") as f:
            json.dump(notes_data, f, indent=2, ensure_ascii=False)
        
        print(f"\n✅ JSON saved: {json_file_path}")
        print(f"📊 Total notes processed: {len(notes_data)}")
        
        # Enhanced summary
        print("\n" + "="*80)
        print("📈 NOTES SUMMARY")
        print("="*80)
        for note_num in sorted(notes_data.keys()):
            note = notes_data[note_num]
            items = len(note["structure"][0]["subcategories"])
            change = note['total_2024'] - note['total_2023']
            change_pct = (change / note['total_2023'] * 100) if note['total_2023'] != 0 else 0
            print(f"Note {note_num:>2}: {note['category_name'][:35]:<35} | Items: {items:>2} | "
                  f"2024: {note['total_2024']:>10.2f} | 2023: {note['total_2023']:>10.2f} | "
                  f"Change: {change:>8.2f} ({change_pct:>6.1f}%)")
        
        return notes_data
    
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return {}

def generate_balance_sheet_report(notes_data):
    """Generate comprehensive Balance Sheet matching the template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"

    # Define styles
    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True, size=10)
    normal_font = Font(size=10)
    bold_font = Font(bold=True, size=10)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))
    top_bottom_border = Border(top=Side(style="thin"), bottom=Side(style="thin"))
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", indent=1)
    right_align = Alignment(horizontal="right", vertical="center")

    # Set column widths
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15

    row = 1

    # Title
    ws.merge_cells("A1:D1")
    ws["A1"] = "Balance Sheet as at March 31, 2024"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align
    ws["A1"].border = top_bottom_border
    row += 1

    # In Lakhs
    ws["C2"] = "In Lakhs"
    ws["C2"].font = normal_font
    ws["C2"].alignment = right_align
    row += 1

    # Headers
    headers = ["", "Notes", "March 31, 2024", "March 31, 2023"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_align
        cell.border = top_bottom_border
    row += 1

    def add_data_row(description, note_ref, val_2024, val_2023, indent=0, is_bold=False, is_section_header=False):
        """Add a data row with proper formatting."""
        nonlocal row
        cell_a = ws.cell(row=row, column=1)
        cell_a.value = description
        cell_a.font = bold_font if (is_bold or is_section_header) else normal_font
        cell_a.alignment = Alignment(horizontal="left", vertical="center", indent=indent)
        if not is_section_header:
            cell_a.border = thin_border
        cell_b = ws.cell(row=row, column=2)
        cell_b.value = note_ref if note_ref else ""
        cell_b.font = normal_font
        cell_b.alignment = center_align
        if not is_section_header:
            cell_b.border = thin_border
        cell_c = ws.cell(row=row, column=3)
        cell_c.value = f"{val_2024:,.2f}" if val_2024 != 0 else ""
        cell_c.font = bold_font if is_bold else normal_font
        cell_c.alignment = right_align
        if not is_section_header:
            cell_c.border = thin_border
        cell_d = ws.cell(row=row, column=4)
        cell_d.value = f"{val_2023:,.2f}" if val_2023 != 0 else ""
        cell_d.font = bold_font if is_bold else normal_font
        cell_d.alignment = right_align
        if not is_section_header:
            cell_d.border = thin_border
        row += 1

    # Equity and Liabilities
    add_data_row("Equity and liabilities", "", 0, 0, is_section_header=True)
    
    # Shareholders' Funds
    add_data_row("Shareholders' funds", "", 0, 0, indent=1, is_section_header=True)
    
    # Share Capital (Note 2)
    share_capital_2024 = notes_data.get("2", {}).get("total_2024", 0.0)
    share_capital_2023 = notes_data.get("2", {}).get("total_2023", 0.0)
    if share_capital_2024 == 0 and share_capital_2023 == 0:
        print("⚠ Warning: No data found for Note 2 (Share capital)")
    add_data_row("Share capital", "2", share_capital_2024, share_capital_2023, indent=2)
    
    # Reserves and Surplus (Note 3)
    reserves_2024 = notes_data.get("3", {}).get("total_2024", 0.0)
    reserves_2023 = notes_data.get("3", {}).get("total_2023", 0.0)
    if reserves_2024 == 0 and reserves_2023 == 0:
        print("⚠ Warning: No data found for Note 3 (Reserves and surplus)")
    add_data_row("Reserves and surplus", "3", reserves_2024, reserves_2023, indent=2)
    
    # Total Shareholders' Funds
    total_equity_2024 = share_capital_2024 + reserves_2024
    total_equity_2023 = share_capital_2023 + reserves_2023
    add_data_row("", "", total_equity_2024, total_equity_2023, indent=1, is_bold=True)
    
    # Non-Current Liabilities
    add_data_row("Non-Current liabilities", "", 0, 0, indent=1, is_section_header=True)
    
    # Long Term Borrowings (Note 4)
    borrowings_2024 = notes_data.get("4", {}).get("total_2024", 0.0)
    borrowings_2023 = notes_data.get("4", {}).get("total_2023", 0.0)
    if borrowings_2024 == 0 and borrowings_2023 == 0:
        print("⚠ Warning: No data found for Note 4 (Long term borrowings)")
    add_data_row("Long term borrowings", "4", borrowings_2024, borrowings_2023, indent=2)
    
    # Deferred Tax Liability (Note 5)
    deferred_tax_2024 = notes_data.get("5", {}).get("total_2024", 0.0)
    deferred_tax_2023 = notes_data.get("5", {}).get("total_2023", 0.0)
    if deferred_tax_2024 == 0 and deferred_tax_2023 == 0:
        print("⚠ Warning: No data found for Note 5 (Deferred tax liability)")
    add_data_row("Deferred Tax Liability (Net)", "5", deferred_tax_2024, deferred_tax_2023, indent=2)
    
    # Total Non-Current Liabilities
    total_non_current_2024 = borrowings_2024 + deferred_tax_2024
    total_non_current_2023 = borrowings_2023 + deferred_tax_2023
    add_data_row("", "", total_non_current_2024, total_non_current_2023, indent=1, is_bold=True)
    
    # Current Liabilities
    add_data_row("Current liabilities", "", 0, 0, indent=1, is_section_header=True)
    
    # Trade Payables (Note 6)
    trade_payables_2024 = notes_data.get("6", {}).get("total_2024", 0.0)
    trade_payables_2023 = notes_data.get("6", {}).get("total_2023", 0.0)
    if trade_payables_2024 == 0 and trade_payables_2023 == 0:
        print("⚠ Warning: No data found for Note 6 (Trade payables)")
    add_data_row("Trade payables", "6", trade_payables_2024, trade_payables_2023, indent=2)
    
    # Other Current Liabilities (Note 7)
    other_liabilities_2024 = notes_data.get("7", {}).get("total_2024", 0.0)
    other_liabilities_2023 = notes_data.get("7", {}).get("total_2023", 0.0)
    if other_liabilities_2024 == 0 and other_liabilities_2023 == 0:
        print("⚠ Warning: No data found for Note 7 (Other current liabilities)")
    add_data_row("Other current liabilities", "7", other_liabilities_2024, other_liabilities_2023, indent=2)
    
    # Short Term Provisions (Note 8)
    provisions_2024 = notes_data.get("8", {}).get("total_2024", 0.0)
    provisions_2023 = notes_data.get("8", {}).get("total_2023", 0.0)
    if provisions_2024 == 0 and provisions_2023 == 0:
        print("⚠ Warning: No data found for Note 8 (Short term provisions)")
    add_data_row("Short term provisions", "8", provisions_2024, provisions_2023, indent=2)
    
    # Total Current Liabilities
    total_current_liab_2024 = trade_payables_2024 + other_liabilities_2024 + provisions_2024
    total_current_liab_2023 = trade_payables_2023 + other_liabilities_2023 + provisions_2023
    add_data_row("", "", total_current_liab_2024, total_current_liab_2023, indent=1, is_bold=True)
    
    # Total Equity and Liabilities
    total_liabilities_2024 = total_equity_2024 + total_non_current_2024 + total_current_liab_2024
    total_liabilities_2023 = total_equity_2023 + total_non_current_2023 + total_current_liab_2023
    add_data_row("TOTAL", "", total_liabilities_2024, total_liabilities_2023, is_bold=True)
    
    # Assets
    add_data_row("Assets", "", 0, 0, is_section_header=True)
    
    # Non-Current Assets
    add_data_row("Non-current assets", "", 0, 0, indent=1, is_section_header=True)
    
    # Fixed Assets (Note 9)
    add_data_row("Fixed assets", "9", 0, 0, indent=2, is_section_header=True)
    
    # Tangible Assets
    tangible_2024 = sum(item["value"] for item in notes_data.get("9", {}).get("structure", [{}])[0].get("subcategories", []) 
                       if not any(x in item["label"].lower() for x in ["intangible", "software"]))
    tangible_2023 = sum(item["previous_value"] for item in notes_data.get("9", {}).get("structure", [{}])[0].get("subcategories", []) 
                       if not any(x in item["label"].lower() for x in ["intangible", "software"]))
    if tangible_2024 == 0 and tangible_2023 == 0:
        print("⚠ Warning: No data found for Tangible assets (Note 9)")
    add_data_row("Tangible assets", "", tangible_2024, tangible_2023, indent=3)
    
    # Intangible Assets
    intangible_2024 = sum(item["value"] for item in notes_data.get("9", {}).get("structure", [{}])[0].get("subcategories", []) 
                         if any(x in item["label"].lower() for x in ["intangible", "software"]))
    intangible_2023 = sum(item["previous_value"] for item in notes_data.get("9", {}).get("structure", [{}])[0].get("subcategories", []) 
                         if any(x in item["label"].lower() for x in ["intangible", "software"]))
    if intangible_2024 == 0 and intangible_2023 == 0:
        print("⚠ Warning: No data found for Intangible assets (Note 9)")
    add_data_row("Intangible assets", "", intangible_2024, intangible_2023, indent=3)
    
    # Capital Work in Progress
    capital_wip_2024, capital_wip_2023 = 0.0, 0.0
    print("⚠ Warning: No data found for Capital Work in Progress")
    add_data_row("Capital Work in Progress", "", capital_wip_2024, capital_wip_2023, indent=3)
    
    # Long Term Loans and Advances (Note 10)
    long_term_loans_2024 = notes_data.get("10", {}).get("total_2024", 0.0)
    long_term_loans_2023 = notes_data.get("10", {}).get("total_2023", 0.0)
    if long_term_loans_2024 == 0 and long_term_loans_2023 == 0:
        print("⚠ Warning: No data found for Note 10 (Long term loans and advances)")
    add_data_row("Long Term Loans and Advances", "10", long_term_loans_2024, long_term_loans_2023, indent=2)
    
    # Total Non-Current Assets
    total_non_current_assets_2024 = tangible_2024 + intangible_2024 + capital_wip_2024 + long_term_loans_2024
    total_non_current_assets_2023 = tangible_2023 + intangible_2023 + capital_wip_2023 + long_term_loans_2023
    add_data_row("", "", total_non_current_assets_2024, total_non_current_assets_2023, indent=1, is_bold=True)
    
    # Current Assets
    add_data_row("Current assets", "", 0, 0, indent=1, is_section_header=True)
    
    # Inventories (Note 11)
    inventories_2024 = notes_data.get("11", {}).get("total_2024", 0.0)
    inventories_2023 = notes_data.get("11", {}).get("total_2023", 0.0)
    if inventories_2024 == 0 and inventories_2023 == 0:
        print("⚠ Warning: No data found for Note 11 (Inventories)")
    add_data_row("Inventories", "11", inventories_2024, inventories_2023, indent=2)
    
    # Trade Receivables (Note 12)
    trade_receivables_2024 = notes_data.get("12", {}).get("total_2024", 0.0)
    trade_receivables_2023 = notes_data.get("12", {}).get("total_2023", 0.0)
    if trade_receivables_2024 == 0 and trade_receivables_2023 == 0:
        print("⚠ Warning: No data found for Note 12 (Trade receivables)")
    add_data_row("Trade receivables", "12", trade_receivables_2024, trade_receivables_2023, indent=2)
    
    # Cash and Bank Balances (Note 13)
    cash_balances_2024 = notes_data.get("13", {}).get("total_2024", 0.0)
    cash_balances_2023 = notes_data.get("13", {}).get("total_2023", 0.0)
    if cash_balances_2024 == 0 and cash_balances_2023 == 0:
        print("⚠ Warning: No data found for Note 13 (Cash and bank balances)")
    add_data_row("Cash and bank balances", "13", cash_balances_2024, cash_balances_2023, indent=2)
    
    # Short Term Loans and Advances (Note 14)
    short_term_loans_2024 = notes_data.get("14", {}).get("total_2024", 0.0)
    short_term_loans_2023 = notes_data.get("14", {}).get("total_2023", 0.0)
    if short_term_loans_2024 == 0 and short_term_loans_2023 == 0:
        print("⚠ Warning: No data found for Note 14 (Short term loans and advances)")
    add_data_row("Short-term loans and advances", "14", short_term_loans_2024, short_term_loans_2023, indent=2)
    
    # Other Current Assets (Note 15)
    other_assets_2024 = notes_data.get("15", {}).get("total_2024", 0.0)
    other_assets_2023 = notes_data.get("15", {}).get("total_2023", 0.0)
    if other_assets_2024 == 0 and other_assets_2023 == 0:
        print("⚠ Warning: No data found for Note 15 (Other current assets)")
    add_data_row("Other current assets", "15", other_assets_2024, other_assets_2023, indent=2)
    
    # Total Current Assets
    total_current_assets_2024 = (inventories_2024 + trade_receivables_2024 + cash_balances_2024 + 
                                short_term_loans_2024 + other_assets_2024)
    total_current_assets_2023 = (inventories_2023 + trade_receivables_2023 + cash_balances_2023 + 
                                short_term_loans_2023 + other_assets_2023)
    add_data_row("", "", total_current_assets_2024, total_current_assets_2023, indent=1, is_bold=True)
    
    # Total Assets
    total_assets_2024 = total_non_current_assets_2024 + total_current_assets_2024
    total_assets_2023 = total_non_current_assets_2023 + total_current_assets_2023
    add_data_row("TOTAL", "", total_assets_2024, total_assets_2023, is_bold=True)
    
    # Footer Notes
    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "The accompanying notes are an integral part of the financial statements"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    row += 1
    
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "As per my report of even date. For and on behalf of the Board of Directors"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    row += 2
    
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "For M/s Siva Parvathi & Associates"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    row += 1
    
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "ICAI Firm registration number: 020872S"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    row += 1
    
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "Chartered Accountants"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    row += 2
    
    ws[f"A{row}"] = "S. Siva Parvathi"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    ws[f"D{row}"] = "Director"
    ws[f"D{row}"].font = normal_font
    ws[f"D{row}"].alignment = right_align
    row += 1
    
    ws[f"A{row}"] = "Proprietor"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    ws[f"D{row}"] = "Director"
    ws[f"D{row}"].font = normal_font
    ws[f"D{row}"].alignment = right_align
    row += 1
    
    ws[f"A{row}"] = "Membership No.:"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    row += 1
    
    ws[f"A{row}"] = "UDIN: 24226087BKEECZ1200"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    row += 1
    
    ws[f"A{row}"] = "Place: Hyderabad"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align
    row += 1
    
    ws[f"A{row}"] = "Date: 04/09/2024"
    ws[f"A{row}"].font = normal_font
    ws[f"A{row}"].alignment = left_align

    # Apply borders
    for r in range(1, row):
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            if cell.value and not any(keyword in str(cell.value).lower() for keyword in ["balance sheet", "in lakhs", "notes"]):
                cell.border = thin_border

    # Save the file
    output_folder = "notestoALL"
    os.makedirs(output_folder, exist_ok=True)
    output_file = os.path.join(output_folder, "outputbs_Sheet.xlsx")
    try:
        wb.save(output_file)
        print(f"\n✅ Balance Sheet generated: {output_file}")
        
        # Print summary
        print("\n" + "="*60)
        print("📊 BALANCE SHEET SUMMARY")
        print("="*60)
        print(f"Total Equity 2024:         ₹{total_equity_2024:>12,.2f} Lakhs")
        print(f"Total Equity 2023:         ₹{total_equity_2023:>12,.2f} Lakhs")
        print(f"Total Non-Current Liab 2024: ₹{total_non_current_2024:>12,.2f} Lakhs")
        print(f"Total Non-Current Liab 2023: ₹{total_non_current_2023:>12,.2f} Lakhs")
        print(f"Total Current Liab 2024:   ₹{total_current_liab_2024:>12,.2f} Lakhs")
        print(f"Total Current Liab 2023:   ₹{total_current_liab_2023:>12,.2f} Lakhs")
        print(f"Total Assets 2024:         ₹{total_assets_2024:>12,.2f} Lakhs")
        print(f"Total Assets 2023:         ₹{total_assets_2023:>12,.2f} Lakhs")
    
    except PermissionError:
        print(f"❌ Permission Error: Cannot save to {output_file}")
        fallback_file = os.path.join(os.path.expanduser("~"), "Desktop", "balance_sheet_fallback.xlsx")
        try:
            wb.save(fallback_file)
            print(f"✅ Balance Sheet saved to: {fallback_file}")
        except Exception as e:
            print(f"❌ Failed to save: {str(e)}")
    except Exception as e:
        print(f"❌ Error saving file: {str(e)}")

def main():
    print("🚀 BALANCE SHEET GENERATOR")
    print("=" * 50)
    
    print("\n📋 STEP 1: Converting Excel to JSON")
    notes_data = load_and_map_excel_notes()
    
    if notes_data:
        print("\n📊 STEP 2: Generating Balance Sheet")
        generate_balance_sheet_report(notes_data)
        print("\n🎉 PROCESS COMPLETED!")
    else:
        print("\n❌ FAILED: No data found")

if __name__ == "__main__":
    main()