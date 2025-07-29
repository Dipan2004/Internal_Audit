import os
import json
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
import requests
from dotenv import load_dotenv

load_dotenv()

try:
    from temp_bl import BalanceSheet as BalanceSheetTemplate
    print("✅ Template imported")
except ImportError:
    print("❌ temp_bl.py not found")
    exit(1)

class BalanceSheetGenerator:
    def __init__(self, api_key: str):
        self.api_key = api_key
        self.base_url = "https://openrouter.ai/api/v1/chat/completions"
        
        # Simplified keyword mapping
        self.keywords = {
            "share_capital": ["share capital", "equity share", "paid up"],
            "reserves_surplus": ["reserves", "surplus", "retained earnings"],
            "long_term_borrowings": ["long term", "borrowings", "debt"],
            "trade_payables": ["trade payables", "payables", "creditors"],
            "other_current_liabilities": ["other current", "accrued"],
            "tangible_assets": ["tangible", "property plant", "fixed assets"],
            "inventories": ["inventories", "stock", "consumables"],
            "trade_receivables": ["trade receivables", "receivables", "debtors"],
            "cash_bank": ["cash", "bank balances", "cash equivalents"],
            "short_term_loans_advances": ["short term", "advances", "prepaid"],
            "other_current_assets": ["other current assets", "accrued income"]
        }

    def safe_float(self, value) -> float:
        if not value or str(value).strip() in ['-', '--', 'None', '']:
            return 0.0
        str_val = re.sub(r'[₹,Rs\.\s\(\)]', '', str(value)).replace('(', '-').replace(')', '')
        try:
            return float(str_val)
        except:
            return 0.0

    def call_api(self, messages, model="mistralai/mixtral-8x7b-instruct"):
        headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json"
        }
        payload = {"model": model, "messages": messages, "temperature": 0.1, "max_tokens": 3000}
        
        try:
            response = requests.post(self.base_url, headers=headers, json=payload, timeout=45)
            return response.json()['choices'][0]['message']['content']
        except:
            return "{\"extracted_items\": []}"

    def extract_data(self, json_data: dict) -> dict:
        print("🔍 Extracting data...")
        items = []
        
        for file_key, file_data in json_data.get("files", {}).items():
            ai_insights = file_data.get("data", {}).get("ai_insights", {})
            
            if isinstance(ai_insights, dict):
                financial_data = ai_insights.get("FinancialData", [])
                for item in financial_data:
                    section = item.get("Section", "")
                    subsection = item.get("Subsection", "")
                    name = f"{section} - {subsection}".strip(" -")
                    
                    val_2024 = self.safe_float(item.get("Value_2024", item.get("Value", 0)))
                    val_2023 = self.safe_float(item.get("Value_2023", 0))
                    
                    if val_2024 or val_2023:
                        items.append({
                            "name": name,
                            "value_2024": val_2024,
                            "value_2023": val_2023,
                            "note": str(item.get("Note", "")).strip()
                        })
        
        # Try AI extraction if direct fails
        if not items:
            try:
                summary = json.dumps({k: v.get("data", {}).get("ai_insights", {}) 
                                   for k, v in json_data.get("files", {}).items()})[:4000]
                
                prompt = f"""Extract balance sheet data from JSON. Return only:
{{"extracted_items": [{{"name": "Share Capital", "value_2024": 542.52, "value_2023": 542.52, "note": "2"}}]}}

Data: {summary}"""
                
                response = self.call_api([{"role": "user", "content": prompt}])
                response = re.sub(r'```(?:json)?\s*', '', response).strip('`').strip()
                
                ai_data = json.loads(response)
                for item in ai_data.get("extracted_items", []):
                    if item.get("name") and (item.get("value_2024") or item.get("value_2023")):
                        items.append({
                            "name": item["name"],
                            "value_2024": self.safe_float(item.get("value_2024", 0)),
                            "value_2023": self.safe_float(item.get("value_2023", 0)),
                            "note": str(item.get("note", "")).strip()
                        })
            except:
                pass
        
        print(f"✅ Found {len(items)} items")
        return {"items": items, "currency": "Lakhs"}

    def match_item(self, name: str):
        name_lower = name.lower()
        for template_key, keywords in self.keywords.items():
            for keyword in keywords:
                if keyword in name_lower:
                    return template_key
        return None

    def populate_template(self, data: dict) -> BalanceSheetTemplate:
        print("📊 Populating template...")
        bs = BalanceSheetTemplate()
        mapped = 0
        
        # Mapping to template structure
        mapping = {
            "share_capital": (bs.equity_liabilities["Shareholders' funds"], "Share capital"),
            "reserves_surplus": (bs.equity_liabilities["Shareholders' funds"], "Reserves and surplus"),
            "long_term_borrowings": (bs.equity_liabilities["Non-Current liabilities"], "Long term borrowings"),
            "trade_payables": (bs.equity_liabilities["Current liabilities"], "Trade payables"),
            "other_current_liabilities": (bs.equity_liabilities["Current liabilities"], "Other current liabilities"),
            "tangible_assets": (bs.assets["Non-current assets"]["Fixed assets"], "Tangible assets"),
            "inventories": (bs.assets["Current assets"], "Inventories"),
            "trade_receivables": (bs.assets["Current assets"], "Trade receivables"),
            "cash_bank": (bs.assets["Current assets"], "Cash and bank balances"),
            "short_term_loans_advances": (bs.assets["Current assets"], "Short-term loans and advances"),
            "other_current_assets": (bs.assets["Current assets"], "Other current assets")
        }
        
        for item in data["items"]:
            match = self.match_item(item["name"])
            if match and match in mapping:
                target_dict, field_name = mapping[match]
                target_dict[field_name].update({
                    "note": item["note"],
                    "2024": item["value_2024"],
                    "2023": item["value_2023"]
                })
                mapped += 1
        
        print(f"📊 Mapped {mapped}/{len(data['items'])} items")
        return bs

    def generate_excel(self, bs: BalanceSheetTemplate, currency: str, output_dir: str = "output"):
        os.makedirs(output_dir, exist_ok=True)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Balance Sheet"
        
        # Set column widths
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        
        row = 1
        
        def add_row(desc, note, val_2024, val_2023, bold=False, indent=0):
            nonlocal row
            ws.cell(row=row, column=1, value="  " * indent + desc)
            ws.cell(row=row, column=2, value=note)
            
            for col, val in [(3, val_2024), (4, val_2023)]:
                cell = ws.cell(row=row, column=col)
                if val: 
                    cell.value = val
                    cell.number_format = '#,##0.00'
                if bold:
                    cell.font = Font(bold=True)
            
            if bold:
                ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1
        
        # Header
        add_row("Balance Sheet", "", 0, 0, True)
        add_row(f"(In {currency})", "", 0, 0)
        add_row("", "Note", "As at Mar 31, 2024", "As at Mar 31, 2023", True)
        
        # Process sections
        def process_section(section_data, title, indent=0):
            add_row(title, "", 0, 0, True, indent)
            total_2024 = total_2023 = 0
            
            for name, data in section_data.items():
                if isinstance(data, dict) and "2024" in data:
                    val_2024 = data.get("2024", 0) or 0
                    val_2023 = data.get("2023", 0) or 0
                    if val_2024 or val_2023:
                        add_row(name, data.get("note", ""), val_2024, val_2023, False, indent+1)
                        total_2024 += val_2024
                        total_2023 += val_2023
                elif isinstance(data, dict):
                    for sub_name, sub_data in data.items():
                        if isinstance(sub_data, dict) and "2024" in sub_data:
                            val_2024 = sub_data.get("2024", 0) or 0
                            val_2023 = sub_data.get("2023", 0) or 0
                            if val_2024 or val_2023:
                                add_row(f"{name} - {sub_name}", sub_data.get("note", ""), 
                                       val_2024, val_2023, False, indent+1)
                                total_2024 += val_2024
                                total_2023 += val_2023
            
            if total_2024 or total_2023:
                add_row(f"Total {title}", "", total_2024, total_2023, True, indent)
            return total_2024, total_2023
        
        # Generate sections
        add_row("EQUITY AND LIABILITIES", "", 0, 0, True)
        eq_2024, eq_2023 = process_section(bs.equity_liabilities["Shareholders' funds"], "Equity", 1)
        ncl_2024, ncl_2023 = process_section(bs.equity_liabilities["Non-Current liabilities"], "Non-Current Liabilities", 1)
        cl_2024, cl_2023 = process_section(bs.equity_liabilities["Current liabilities"], "Current Liabilities", 1)
        
        total_eq_2024 = eq_2024 + ncl_2024 + cl_2024
        total_eq_2023 = eq_2023 + ncl_2023 + cl_2023
        add_row("TOTAL EQUITY & LIABILITIES", "", total_eq_2024, total_eq_2023, True)
        
        add_row("", "", 0, 0)
        add_row("ASSETS", "", 0, 0, True)
        nca_2024, nca_2023 = process_section(bs.assets["Non-current assets"], "Non-Current Assets", 1)
        ca_2024, ca_2023 = process_section(bs.assets["Current assets"], "Current Assets", 1)
        
        total_assets_2024 = nca_2024 + ca_2024
        total_assets_2023 = nca_2023 + ca_2023
        add_row("TOTAL ASSETS", "", total_assets_2024, total_assets_2023, True)
        
        # Save
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = os.path.join(output_dir, f"balance_sheet_{timestamp}.xlsx")
        wb.save(output_file)
        
        # Summary
        balanced = abs(total_assets_2024 - total_eq_2024) < 0.01
        print(f"✅ Generated: {output_file}")
        print(f"📊 Assets: ₹{total_assets_2024:,.2f} | Equity: ₹{total_eq_2024:,.2f}")
        print(f"✅ {'Balanced' if balanced else 'Not Balanced'}")
        
        return output_file

    def process(self, input_file: str, output_dir: str = "output"):
        try:
            print(f"📂 Processing: {input_file}")
            with open(input_file, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
            
            data = self.extract_data(json_data)
            if not data["items"]:
                print("❌ No data extracted")
                return None
            
            bs = self.populate_template(data)
            return self.generate_excel(bs, data["currency"], output_dir)
            
        except Exception as e:
            print(f"❌ Error: {e}")
            return None

def main():
    print("🚀 COMPACT BALANCE SHEET GENERATOR")
    
    api_key = os.getenv("OPENROUTER_API_KEY")
    input_file = os.getenv("INPUT_FILE", "output/ai_enhanced_data.json")
    
    if not api_key:
        print("❌ Missing OPENROUTER_API_KEY")
        return
    
    if not os.path.exists(input_file):
        print(f"❌ File not found: {input_file}")
        return
    
    generator = BalanceSheetGenerator(api_key)
    result = generator.process(input_file)
    
    if result:
        print(f"✅ SUCCESS: {result}")
    else:
        print("❌ FAILED")

if __name__ == "__main__":
    main()