import os
import json
import logging
from typing import Any, Dict, List, Optional
from pydantic import BaseModel, ValidationError
from pydantic_settings import BaseSettings
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class Settings(BaseSettings):
	"""Application settings loaded from environment variables or .env file."""
	input_file: str = "data/output2/notes_output.json"
	output_folder: str = "data/output3"
	output_file: str = "data/final_notes_output.xlsx"

settings = Settings()

class BreakdownItem(BaseModel):
	description: str
	amount: float
	amount_lakhs: Optional[float] = None

class MatchedAccount(BaseModel):
	account: str
	amount: float
	amount_lakhs: Optional[float] = None
	group: Optional[str] = None

class NoteData(BaseModel):
	note_number: Optional[str] = None
	note_title: Optional[str] = None
	full_title: Optional[str] = None
	table_data: Optional[List[Dict[str, Any]]] = []
	breakdown: Optional[Dict[str, BreakdownItem]] = {}
	matched_accounts: Optional[List[MatchedAccount]] = []
	total_amount: Optional[float] = None
	total_amount_lakhs: Optional[float] = None
	matched_accounts_count: Optional[int] = None
	comparative_data: Optional[Dict[str, Any]] = {}
	notes_and_disclosures: Optional[List[str]] = []
	markdown_content: Optional[str] = ""

def create_output_folder(folder_path: str) -> None:
	"""Create output folder if it doesn't exist."""
	if not os.path.exists(folder_path):
		os.makedirs(folder_path)
		logger.info(f"Created folder: {folder_path}")

def read_json_file(file_path: str) -> Optional[Dict[str, Any]]:
	"""Read and parse JSON file."""
	try:
		with open(file_path, 'r', encoding='utf-8') as file:
			data = json.load(file)
		logger.info(f"Successfully read JSON file: {file_path}")
		return data
	except FileNotFoundError:
		logger.error(f"File '{file_path}' not found.")
		return None
	except json.JSONDecodeError as e:
		logger.error(f"Invalid JSON format in '{file_path}': {e}")
		return None
	except Exception as e:
		logger.error(f"Error reading file '{file_path}': {e}")
		return None

def normalize_llm_note_json(llm_json: Dict[str, Any]) -> Dict[str, Any]:
	"""
	Convert LLM note JSON (single note, custom structure) to the standard notes_output.json format.
	"""
	if "note_number" in llm_json or "full_title" in llm_json or "table_data" in llm_json:
		return llm_json

	normalized = {
		"note_number": llm_json.get("metadata", {}).get("note_number", ""),
		"note_title": llm_json.get("title", ""),
		"full_title": llm_json.get("full_title", ""),
		"table_data": [],
		"breakdown": {},
		"matched_accounts": [],
		"total_amount": None,
		"total_amount_lakhs": None,
		"matched_accounts_count": None,
		"comparative_data": {},
		"notes_and_disclosures": [],
		"markdown_content": "",
	}
	if "structure" in llm_json:
		for item in llm_json["structure"]:
			if "category" in item and "subcategories" in item:
				for sub in item["subcategories"]:
					row = {
						"particulars": sub.get("label", ""),
						"current_year": sub.get("value", ""),
						"previous_year": ""
					}
					normalized["table_data"].append(row)
	return normalized

def create_financial_table_sheet(workbook: Workbook, sheet_name: str, note_data: Dict[str, Any]) -> None:
	"""Create a properly formatted financial table sheet."""
	ws = workbook.create_sheet(title=sheet_name)
	header_font = Font(bold=True, color="FFFFFF")
	header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
	bold_font = Font(bold=True)
	center_alignment = Alignment(horizontal="center", vertical="center")
	right_alignment = Alignment(horizontal="right", vertical="center")
	thin_border = Border(
		left=Side(style='thin'),
		right=Side(style='thin'),
		top=Side(style='thin'),
		bottom=Side(style='thin')
	)
	current_row = 1

	# Add Note Title
	note_title = note_data.get('full_title', note_data.get('note_title', 'Note'))
	ws.cell(row=current_row, column=1, value=note_title)
	ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
	current_row += 2

	# Process table_data if available
	if 'table_data' in note_data and note_data['table_data']:
		table_data = note_data['table_data']
		df = pd.DataFrame(table_data)
		for col_num, column_name in enumerate(df.columns, 1):
			cell = ws.cell(row=current_row, column=col_num, value=column_name.replace('_', ' ').title())
			cell.font = header_font
			cell.fill = header_fill
			cell.alignment = center_alignment
			cell.border = thin_border
		current_row += 1
		for _, row in df.iterrows():
			for col_num, value in enumerate(row, 1):
				cell = ws.cell(row=current_row, column=col_num, value=value)
				cell.border = thin_border
				if col_num > 1:
					cell.alignment = right_alignment
				if isinstance(value, str) and ('**' in value or 'Total' in value or 'Particulars' in value):
					cell.font = bold_font
					cell.value = value.replace('**', '')
			current_row += 1
		current_row += 1

	# Add breakdown information if available
	if 'breakdown' in note_data and note_data['breakdown']:
		ws.cell(row=current_row, column=1, value="Breakdown Details:")
		ws.cell(row=current_row, column=1).font = bold_font
		current_row += 1
		ws.cell(row=current_row, column=1, value="Description")
		ws.cell(row=current_row, column=2, value="Amount")
		ws.cell(row=current_row, column=3, value="Amount (Lakhs)")
		for col in range(1, 4):
			cell = ws.cell(row=current_row, column=col)
			cell.font = header_font
			cell.fill = header_fill
			cell.alignment = center_alignment
			cell.border = thin_border
		current_row += 1
		for key, value in note_data['breakdown'].items():
			if isinstance(value, dict):
				desc = value.get('description', key)
				amount = value.get('amount', 0)
				amount_lakhs = value.get('amount_lakhs', 0)
				ws.cell(row=current_row, column=1, value=desc).border = thin_border
				ws.cell(row=current_row, column=2, value=amount).border = thin_border
				ws.cell(row=current_row, column=3, value=amount_lakhs).border = thin_border
				ws.cell(row=current_row, column=2).alignment = right_alignment
				ws.cell(row=current_row, column=3).alignment = right_alignment
				current_row += 1
		current_row += 1

	# Add matched accounts if available
	if 'matched_accounts' in note_data and note_data['matched_accounts']:
		ws.cell(row=current_row, column=1, value="Account-wise Breakdown:")
		ws.cell(row=current_row, column=1).font = bold_font
		current_row += 1
		headers = ["Account", "Amount", "Amount (Lakhs)", "Group"]
		for col_num, header in enumerate(headers, 1):
			cell = ws.cell(row=current_row, column=col_num, value=header)
			cell.font = header_font
			cell.fill = header_fill
			cell.alignment = center_alignment
			cell.border = thin_border
		current_row += 1
		for account in note_data['matched_accounts']:
			ws.cell(row=current_row, column=1, value=account.get('account', '')).border = thin_border
			ws.cell(row=current_row, column=2, value=account.get('amount', 0)).border = thin_border
			ws.cell(row=current_row, column=3, value=account.get('amount_lakhs', 0)).border = thin_border
			ws.cell(row=current_row, column=4, value=account.get('group', '')).border = thin_border
			ws.cell(row=current_row, column=2).alignment = right_alignment
			ws.cell(row=current_row, column=3).alignment = right_alignment
			current_row += 1
		current_row += 1

	# Add summary information
	if 'total_amount' in note_data:
		ws.cell(row=current_row, column=1, value="Summary:")
		ws.cell(row=current_row, column=1).font = bold_font
		current_row += 1
		ws.cell(row=current_row, column=1, value="Total Amount:")
		ws.cell(row=current_row, column=2, value=note_data.get('total_amount', 0))
		ws.cell(row=current_row, column=2).alignment = right_alignment
		current_row += 1
		ws.cell(row=current_row, column=1, value="Total Amount (Lakhs):")
		ws.cell(row=current_row, column=2, value=note_data.get('total_amount_lakhs', 0))
		ws.cell(row=current_row, column=2).alignment = right_alignment
		current_row += 1
		ws.cell(row=current_row, column=1, value="Matched Accounts Count:")
		ws.cell(row=current_row, column=2, value=note_data.get('matched_accounts_count', 0))
		ws.cell(row=current_row, column=2).alignment = right_alignment
		current_row += 1

	# Auto-adjust column widths
	for column in ws.columns:
		max_length = 0
		column_letter = get_column_letter(column[0].column)
		for cell in column:
			try:
				if len(str(cell.value)) > max_length:
					max_length = len(str(cell.value))
			except Exception:
				pass
		adjusted_width = min(max_length + 2, 60)
		ws.column_dimensions[column_letter].width = adjusted_width

def convert_json_to_excel(input_file: str, output_file: str) -> bool:
	"""Main function to convert JSON to Excel."""
	json_data = read_json_file(input_file)
	if json_data is None:
		return False

	# Normalize if needed
	if isinstance(json_data, dict) and "notes" not in json_data:
		normalized_note = normalize_llm_note_json(json_data)
		json_data = {"notes": [normalized_note]}
	elif isinstance(json_data, list):
		json_data = {"notes": json_data}

	workbook = Workbook()
	default_sheet = workbook.active
	workbook.remove(default_sheet)

	if 'notes' in json_data:
		notes_data = json_data['notes']
		for note in notes_data:
			try:
				validated_note = NoteData(**note)
			except ValidationError as ve:
				logger.warning(f"Validation error for note: {ve}")
				validated_note = note  # fallback to raw dict
			note_title = note.get('full_title', note.get('note_title', f"Note {note.get('note_number', '')}"))
			clean_sheet_name = str(note_title).replace('/', '_').replace('\\', '_').replace('*', '_')
			clean_sheet_name = clean_sheet_name.replace('?', '_').replace('[', '_').replace(']', '_')
			clean_sheet_name = clean_sheet_name[:31]
			logger.info(f"Processing: {clean_sheet_name}")
			create_financial_table_sheet(workbook, clean_sheet_name, note)
	else:
		for note_key, note_data in json_data.items():
			clean_sheet_name = str(note_key).replace('/', '_').replace('\\', '_').replace('*', '_')
			clean_sheet_name = clean_sheet_name.replace('?', '_').replace('[', '_').replace(']', '_')
			clean_sheet_name = clean_sheet_name[:31]
			logger.info(f"Processing: {clean_sheet_name}")
			if isinstance(note_data, dict):
				create_financial_table_sheet(workbook, clean_sheet_name, note_data)
			else:
				simple_data = {"value": note_data}
				create_financial_table_sheet(workbook, clean_sheet_name, simple_data)

	try:
		workbook.save(output_file)
		logger.info(f"Successfully saved Excel file: {output_file}")
		return True
	except Exception as e:
		logger.error(f"Error saving Excel file: {e}")
		return False

def json_to_xlsx(input_json: str, output_xlsx: str) -> None:
	"""
	Convert the given JSON file to Excel using the existing logic.
	"""
	convert_json_to_excel(input_json, output_xlsx)

def main() -> None:
	"""Main execution function."""
	input_file = settings.input_file
	output_folder = settings.output_folder
	output_file = os.path.join(output_folder, settings.output_file)
	create_output_folder(output_folder)

	if not os.path.exists(input_file):
		logger.error(f"Input file '{input_file}' not found. Please ensure the file exists in the correct location.")
		return

	success = convert_json_to_excel(input_file, output_file)

	if success:
		logger.info("=" * 50)
		logger.info("CONVERSION COMPLETED SUCCESSFULLY!")
		logger.info("=" * 50)
		logger.info(f"Input file: {input_file}")
		logger.info(f"Output file: {output_file}")
		logger.info("The Excel file has been created with:")
		logger.info("- Each note as a separate sheet")
		logger.info("- Proper financial table formatting")
		logger.info("- Table data displayed in tabular format")
		logger.info("- Breakdown and account details included")
		logger.info("- Professional styling and formatting")
	else:
		logger.error("=" * 50)
		logger.error("CONVERSION FAILED!")
		logger.error("=" * 50)
		logger.error("Please check the error messages above.")

if __name__ == "__main__":
	main()
