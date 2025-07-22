import os
import json
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side, Alignment
import re

def safe_float_conversion(value):
    """Safely convert value to float, handling various formats."""
    if value is None:
        return 0.0
    
    # Convert to string and clean up
    str_val = str(value).strip()
    
    # Handle empty or dash values
    if not str_val or str_val in ['-', '--', '']:
        return 0.0
    
    # Remove common formatting
    str_val = str_val.replace(',', '').replace('₹', '').replace('Rs.', '')
    str_val = str_val.replace('(', '').replace(')', '')
    
    # Handle parentheses as negative (accounting format)
    is_negative = '(' in str(value) and ')' in str(value)
    
    try:
        result = float(str_val)
        return -result if is_negative else result
    except (ValueError, TypeError):
        print(f"Warning: Could not convert '{value}' to float, using 0.0")
        return 0.0

def find_year_columns(ws, max_search_rows=15):
    """Find columns containing year data (2024, 2023)."""
    year_2024_col = None
    year_2023_col = None
    
    print("Searching for year columns...")
    
    for row_idx in range(1, min(max_search_rows + 1, ws.max_row + 1)):
        for col_idx in range(1, ws.max_column + 1):
            cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "").strip()
            
            if "2024" in cell_value and year_2024_col is None:
                year_2024_col = col_idx
                print(f"Found 2024 column at column {col_idx} (row {row_idx})")
            elif "2023" in cell_value and year_2023_col is None:
                year_2023_col = col_idx
                print(f"Found 2023 column at column {col_idx} (row {row_idx})")
    
    # If not found, try to find by position (common patterns)
    if year_2024_col is None or year_2023_col is None:
        print("Year columns not found in headers, trying positional detection...")
        # Look for numeric data patterns
        for row_idx in range(5, min(20, ws.max_row + 1)):
            row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1)]
            
            for col_idx, value in enumerate(row_values, 1):
                if value is not None and str(value).strip():
                    try:
                        float_val = safe_float_conversion(value)
                        if float_val > 0:  # Found numeric data
                            if year_2024_col is None and col_idx >= 3:
                                year_2024_col = col_idx
                            elif year_2023_col is None and col_idx > year_2024_col:
                                year_2023_col = col_idx
                                break
                    except:
                        continue
            
            if year_2024_col and year_2023_col:
                break
    
    # Default fallback positions
    if year_2024_col is None:
        year_2024_col = 3
        print(f"Using default 2024 column: {year_2024_col}")
    if year_2023_col is None:
        year_2023_col = 4
        print(f"Using default 2023 column: {year_2023_col}")
    
    return year_2024_col, year_2023_col

def load_and_map_excel_notes(file_path="data/notes_pnl2.xlsx"):
    """Load notes data from Excel file with improved numeric handling."""
    try:
        wb = load_workbook(file_path, data_only=True)  # data_only=True to get calculated values
        ws = wb.active
        notes_data = {}
        
        print("Reading Excel file...")
        print(f"Sheet has {ws.max_row} rows and {ws.max_column} columns")
        
        # Find year columns
        year_2024_col, year_2023_col = find_year_columns(ws)
        print(f"Using columns - 2024: {year_2024_col}, 2023: {year_2023_col}")
        
        # Convert all rows to list for processing
        all_rows = []
        for row_idx in range(1, ws.max_row + 1):
            row_data = []
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_data.append(cell_value)
            all_rows.append(row_data)
        
        i = 0
        while i < len(all_rows):
            row = all_rows[i]
            
            if not row or not row[0]:
                i += 1
                continue
                
            cell_value = str(row[0]).strip()
            
            # Look for note patterns (16., 17., ... 28.)
            note_match = re.match(r'^(\d{2})\.?\s*(.+)', cell_value)
            if note_match and int(note_match.group(1)) in range(16, 29):
                note_num = note_match.group(1)
                category_name = note_match.group(2).strip()
                
                print(f"\n📋 Processing Note {note_num}: {category_name}")
                
                # Initialize note structure
                notes_data[note_num] = {
                    "category_name": category_name,
                    "structure": [{
                        "category": category_name,
                        "subcategories": []
                    }]
                }
                
                # Process subsequent rows
                i += 1
                line_count = 0
                
                while i < len(all_rows):
                    data_row = all_rows[i]
                    
                    if not data_row or len(data_row) == 0:
                        i += 1
                        continue
                    
                    data_cell_value = str(data_row[0] or "").strip()
                    
                    # Check if we've hit another note
                    next_note_match = re.match(r'^(\d{2})\.?\s*(.+)', data_cell_value)
                    if next_note_match and int(next_note_match.group(1)) in range(16, 29):
                        i -= 1  # Step back to process this note
                        break
                    
                    # Skip obvious header/formatting rows
                    skip_patterns = [
                        "in lakhs", "march 31", "year ended", "notes", "particulars",
                        "amount", "total", "subtotal", "details", "description"
                    ]
                    
                    if (not data_cell_value or 
                        len(data_cell_value) <= 1 or
                        any(pattern in data_cell_value.lower() for pattern in skip_patterns)):
                        i += 1
                        continue
                    
                    # Extract values with improved error handling
                    value_2024 = 0.0
                    value_2023 = 0.0
                    
                    # Get 2024 value
                    if year_2024_col <= len(data_row):
                        raw_2024 = data_row[year_2024_col - 1]
                        value_2024 = safe_float_conversion(raw_2024)
                    
                    # Get 2023 value
                    if year_2023_col <= len(data_row):
                        raw_2023 = data_row[year_2023_col - 1]
                        value_2023 = safe_float_conversion(raw_2023)
                    
                    # Only add meaningful entries (has text and/or values)
                    if (data_cell_value and len(data_cell_value.strip()) > 2 and 
                        (value_2024 != 0.0 or value_2023 != 0.0 or 
                         any(char.isalpha() for char in data_cell_value))):
                        
                        subcategory = {
                            "label": data_cell_value,
                            "value": value_2024,
                            "previous_value": value_2023,
                            "change": value_2024 - value_2023,
                            "change_percent": ((value_2024 - value_2023) / value_2023 * 100) if value_2023 != 0 else 0
                        }
                        notes_data[note_num]["structure"][0]["subcategories"].append(subcategory)
                        
                        line_count += 1
                        print(f"  ✓ {data_cell_value[:35]:<35} | 2024: {value_2024:>10.2f} | 2023: {value_2023:>10.2f}")
                    
                    i += 1
                
                print(f"  📊 Added {line_count} line items for Note {note_num}")
                
                # Calculate totals for this note
                total_2024 = sum(s["value"] for s in notes_data[note_num]["structure"][0]["subcategories"])
                total_2023 = sum(s["previous_value"] for s in notes_data[note_num]["structure"][0]["subcategories"])
                notes_data[note_num]["total_2024"] = total_2024
                notes_data[note_num]["total_2023"] = total_2023
                notes_data[note_num]["total_change"] = total_2024 - total_2023
                
            i += 1

        # Create output directory and save JSON
        json_file_path = "data/pnl_notes.json"
        os.makedirs(os.path.dirname(json_file_path), exist_ok=True)
        
        with open(json_file_path, "w", encoding="utf-8") as f:
            json.dump(notes_data, f, indent=2, ensure_ascii=False)
        
        print(f"\n✅ JSON file saved: {json_file_path}")
        print(f"📊 Total notes mapped: {len(notes_data)}")
        
        # Print detailed summary
        print("\n" + "="*80)
        print("📈 NOTES SUMMARY")
        print("="*80)
        for note_num in sorted(notes_data.keys(), key=int):
            note_info = notes_data[note_num]
            subcat_count = len(note_info["structure"][0]["subcategories"])
            total_2024 = note_info.get("total_2024", 0)
            total_2023 = note_info.get("total_2023", 0)
            change = note_info.get("total_change", 0)
            
            print(f"Note {note_num}: {note_info['category_name'][:40]:<40}")
            print(f"         Items: {subcat_count:<3} | 2024: {total_2024:>12.2f} | 2023: {total_2023:>12.2f} | Change: {change:>12.2f}")
        
        return notes_data
        
    except FileNotFoundError:
        print(f"❌ Error: Excel file {file_path} not found.")
        print("Please ensure the file exists in the specified location.")
        return {}
    except Exception as e:
        print(f"❌ Error loading Excel file: {str(e)}")
        import traceback
        traceback.print_exc()
        return {}

def load_note_data(file_path="data/pnl_notes.json"):
    """Load note data from JSON file."""
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"❌ Error: JSON file {file_path} not found.")
        return {}
    except json.JSONDecodeError:
        print(f"❌ Error: Invalid JSON in {file_path}")
        return {}

def calculate_note_total(note_data, year="2024"):
    """Calculate total for a note."""
    if not note_data or "structure" not in note_data:
        return 0.0
    
    total = 0.0
    for category in note_data["structure"]:
        for subcat in category.get("subcategories", []):
            try:
                value = subcat.get("value" if year == "2024" else "previous_value", 0.0)
                total += float(value)
            except (ValueError, TypeError):
                continue
    return total

def format_currency(value):
    """Format currency value."""
    if isinstance(value, (int, float)) and value != 0:
        return f"{value:,.2f}"
    return "-"

def generate_comprehensive_pnl_report(notes_data):
    """Generate comprehensive P&L report in Excel format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Profit and Loss Statement"

    # Styles
    bold_font = Font(bold=True, size=12)
    header_font = Font(bold=True, size=14)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))
    center_align = Alignment(horizontal="center")
    left_align = Alignment(horizontal="left")
    right_align = Alignment(horizontal="right")

    # Column widths
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    # Header
    ws["A1"] = "COMPREHENSIVE PROFIT AND LOSS STATEMENT"
    ws["A1"].font = header_font
    ws.merge_cells("A1:E1")
    ws["A1"].alignment = center_align

    ws["A2"] = "For the year ended March 31, 2024"
    ws["A2"].font = bold_font
    ws.merge_cells("A2:E2")
    ws["A2"].alignment = center_align

    ws["A3"] = "(All amounts in Lakhs)"
    ws.merge_cells("A3:E3")
    ws["A3"].alignment = right_align

    # Table headers
    headers = ["Particulars", "Notes", "Year ended March 31, 2024", "Year ended March 31, 2023", "Change"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = center_align

    row = 6

    # INCOME SECTION
    ws.cell(row=row, column=1).value = "INCOME"
    ws.cell(row=row, column=1).font = bold_font
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = thin_border
    row += 1

    # Revenue and Other Income
    total_income_2024 = 0
    total_income_2023 = 0

    # Process income notes (typically 16, 17)
    income_notes = ["16", "17"]
    income_labels = {
        "16": "Revenue from operations (net)",
        "17": "Other income"
    }

    for note_num in income_notes:
        if note_num in notes_data:
            income_2024 = notes_data[note_num].get("total_2024", 0)
            income_2023 = notes_data[note_num].get("total_2023", 0)
            change = income_2024 - income_2023
            
            total_income_2024 += income_2024
            total_income_2023 += income_2023
            
            ws.cell(row=row, column=1).value = income_labels.get(note_num, f"Income Note {note_num}")
            ws.cell(row=row, column=2).value = note_num
            ws.cell(row=row, column=3).value = format_currency(income_2024)
            ws.cell(row=row, column=4).value = format_currency(income_2023)
            ws.cell(row=row, column=5).value = format_currency(change)
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = thin_border
            row += 1

    # Total Income
    ws.cell(row=row, column=1).value = "Total Income (I)"
    ws.cell(row=row, column=1).font = bold_font
    ws.cell(row=row, column=3).value = format_currency(total_income_2024)
    ws.cell(row=row, column=4).value = format_currency(total_income_2023)
    ws.cell(row=row, column=5).value = format_currency(total_income_2024 - total_income_2023)
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=col).font = bold_font
    row += 2

    # EXPENSES SECTION
    ws.cell(row=row, column=1).value = "EXPENSES"
    ws.cell(row=row, column=1).font = bold_font
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = thin_border
    row += 1

    # Process all expense notes
    expense_notes_labels = {
        "18": "Cost of materials consumed",
        "19": "Employee benefit expense",
        "20": "Other expenses",
        "21": "Depreciation and amortisation expense",
        "22": "Impairment losses",
        "23": "Finance costs",
        "24": "Tax expense",
        "25": "Exceptional items",
        "26": "Discontinued operations",
        "27": "Prior period items",
        "28": "Other comprehensive income"
    }

    total_expenses_2024 = 0
    total_expenses_2023 = 0

    # Process expense notes (18-28)
    for note_num in sorted([k for k in notes_data.keys() if int(k) >= 18]):
        note_data = notes_data[note_num]
        expense_2024 = note_data.get("total_2024", 0)
        expense_2023 = note_data.get("total_2023", 0)
        change = expense_2024 - expense_2023
        
        # Only add to total expenses if it's actually an expense (positive values typically)
        if int(note_num) <= 25:  # Typical expense notes
            total_expenses_2024 += expense_2024
            total_expenses_2023 += expense_2023
        
        label = expense_notes_labels.get(note_num, note_data.get("category_name", f"Note {note_num}"))
        
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=2).value = note_num
        ws.cell(row=row, column=3).value = format_currency(expense_2024)
        ws.cell(row=row, column=4).value = format_currency(expense_2023)
        ws.cell(row=row, column=5).value = format_currency(change)
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = thin_border
        row += 1

    # Total Expenses
    ws.cell(row=row, column=1).value = "Total Expenses (II)"
    ws.cell(row=row, column=1).font = bold_font
    ws.cell(row=row, column=3).value = format_currency(total_expenses_2024)
    ws.cell(row=row, column=4).value = format_currency(total_expenses_2023)
    ws.cell(row=row, column=5).value = format_currency(total_expenses_2024 - total_expenses_2023)
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=col).font = bold_font
    row += 2

    # Profit Before Tax
    profit_before_tax_2024 = total_income_2024 - total_expenses_2024
    profit_before_tax_2023 = total_income_2023 - total_expenses_2023
    
    ws.cell(row=row, column=1).value = "Profit/(Loss) before Tax (I) - (II)"
    ws.cell(row=row, column=1).font = bold_font
    ws.cell(row=row, column=3).value = format_currency(profit_before_tax_2024)
    ws.cell(row=row, column=4).value = format_currency(profit_before_tax_2023)
    ws.cell(row=row, column=5).value = format_currency(profit_before_tax_2024 - profit_before_tax_2023)
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = thin_border
        ws.cell(row=row, column=col).font = bold_font

    # Align columns
    for r in range(5, row + 1):
        ws.cell(row=r, column=1).alignment = left_align
        for c in range(2, 6):
            ws.cell(row=r, column=c).alignment = right_align

    # Save file
    output_file = "comprehensive_pnl_report.xlsx"
    try:
        wb.save(output_file)
        print(f"\n✅ Comprehensive P&L report generated: {output_file}")
        
        # Print detailed summary
        print(f"\n" + "="*80)
        print(f"📊 COMPREHENSIVE FINANCIAL SUMMARY")
        print(f"="*80)
        print(f"Total Income 2024:        ₹{format_currency(total_income_2024):>15} Lakhs")
        print(f"Total Income 2023:        ₹{format_currency(total_income_2023):>15} Lakhs")
        print(f"Income Change:            ₹{format_currency(total_income_2024 - total_income_2023):>15} Lakhs")
        print(f"")
        print(f"Total Expenses 2024:      ₹{format_currency(total_expenses_2024):>15} Lakhs")
        print(f"Total Expenses 2023:      ₹{format_currency(total_expenses_2023):>15} Lakhs")
        print(f"Expenses Change:          ₹{format_currency(total_expenses_2024 - total_expenses_2023):>15} Lakhs")
        print(f"")
        print(f"Profit Before Tax 2024:   ₹{format_currency(profit_before_tax_2024):>15} Lakhs")
        print(f"Profit Before Tax 2023:   ₹{format_currency(profit_before_tax_2023):>15} Lakhs")
        print(f"Profit Change:            ₹{format_currency(profit_before_tax_2024 - profit_before_tax_2023):>15} Lakhs")
        
        if total_income_2023 > 0:
            income_growth = ((total_income_2024 - total_income_2023) / total_income_2023) * 100
            print(f"Income Growth:            {income_growth:>18.2f}%")
        
        print(f"="*80)
        
    except Exception as e:
        print(f"❌ Error saving Excel file: {str(e)}")

if __name__ == "__main__":
    print("🚀 IMPROVED P&L GENERATOR STARTING...")
    print("=" * 80)
    
    print("\n📋 STEP 1: Converting Excel to JSON with improved numeric handling")
    notes_data = load_and_map_excel_notes()
    
    if notes_data:
        print("\n📊 STEP 2: Generating Comprehensive P&L Report")
        generate_comprehensive_pnl_report(notes_data)
        print("\n🎉 PROCESS COMPLETED SUCCESSFULLY!")
        print("✅ Check the 'comprehensive_pnl_report.xlsx' file for the complete report")
    else:
        print("\n❌ PROCESS FAILED: No data found.")
        print("Please check:")
        print("  1. Excel file exists at 'data/notes_pnl.xlsx'")
        print("  2. File contains notes numbered 16-28")
        print("  3. Data has proper year columns (2024, 2023)")